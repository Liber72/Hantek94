"""
Mô-đun tạo file Excel bảng cân đối kế toán
Balance Sheet Excel Generator Module

Tạo file Excel với cấu trúc bảng cân đối kế toán chuẩn theo quy định Việt Nam
Creates Excel file with standard balance sheet structure according to Vietnamese regulations

Tuân thủ: Thông tư 200/2014/TT-BTC về chế độ kế toán doanh nghiệp
Compliance: Circular 200/2014/TT-BTC on enterprise accounting regime
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

from data_source import FinancialDataSource

class BalanceSheetGenerator:
    """Lớp tạo file Excel bảng cân đối kế toán"""
    
    def __init__(self, output_directory="output"):
        """
        Khởi tạo generator
        
        Args:
            output_directory (str): Thư mục lưu file output
        """
        self.output_directory = output_directory
        self.data_source = FinancialDataSource()
        
        # Tạo thư mục output nếu chưa tồn tại
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
        
        # Định nghĩa style cho Excel
        self.styles = self._define_styles()
    
    def _define_styles(self):
        """Định nghĩa các style cho Excel"""
        styles = {}
        
        # Font cho tiêu đề chính
        styles['title_font'] = Font(
            name='Times New Roman',
            size=16,
            bold=True
        )
        
        # Font cho tiêu đề phụ
        styles['subtitle_font'] = Font(
            name='Times New Roman',
            size=12,
            bold=True
        )
        
        # Font cho nội dung
        styles['content_font'] = Font(
            name='Times New Roman',
            size=11
        )
        
        # Font cho số liệu
        styles['number_font'] = Font(
            name='Times New Roman',
            size=11
        )
        
        # Alignment giữa
        styles['center_alignment'] = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # Alignment trái
        styles['left_alignment'] = Alignment(
            horizontal='left',
            vertical='center'
        )
        
        # Alignment phải (cho số)
        styles['right_alignment'] = Alignment(
            horizontal='right',
            vertical='center'
        )
        
        # Border
        thin_side = Side(border_style="thin", color="000000")
        styles['thin_border'] = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # Background cho header
        styles['header_fill'] = PatternFill(
            start_color="E6E6FA",
            end_color="E6E6FA",
            fill_type="solid"
        )
        
        # Background cho tổng cộng
        styles['total_fill'] = PatternFill(
            start_color="F0F8FF",
            end_color="F0F8FF",
            fill_type="solid"
        )
        
        return styles
    
    def _apply_cell_style(self, cell, style_name, value=None):
        """Áp dụng style cho cell"""
        if value is not None:
            cell.value = value
        
        if style_name == 'title':
            cell.font = self.styles['title_font']
            cell.alignment = self.styles['center_alignment']
        elif style_name == 'subtitle':
            cell.font = self.styles['subtitle_font']
            cell.alignment = self.styles['center_alignment']
        elif style_name == 'header':
            cell.font = self.styles['subtitle_font']
            cell.alignment = self.styles['center_alignment']
            cell.border = self.styles['thin_border']
            cell.fill = self.styles['header_fill']
        elif style_name == 'content':
            cell.font = self.styles['content_font']
            cell.alignment = self.styles['left_alignment']
            cell.border = self.styles['thin_border']
        elif style_name == 'number':
            cell.font = self.styles['number_font']
            cell.alignment = self.styles['right_alignment']
            cell.border = self.styles['thin_border']
            cell.number_format = '#,##0'
        elif style_name == 'total':
            cell.font = self.styles['subtitle_font']
            cell.alignment = self.styles['right_alignment']
            cell.border = self.styles['thin_border']
            cell.fill = self.styles['total_fill']
            cell.number_format = '#,##0'
    
    def create_balance_sheet(self, filename=None):
        """
        Tạo file Excel bảng cân đối kế toán
        
        Args:
            filename (str): Tên file output (optional)
            
        Returns:
            str: Đường dẫn file đã tạo
        """
        
        # Lấy dữ liệu
        data = self.data_source.get_balance_sheet_data()
        
        # Tạo workbook và worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng Cân Đối Kế Toán"
        
        # Tạo header
        self._create_header(ws, data['company_info'])
        
        # Tạo bảng cân đối kế toán
        current_row = self._create_balance_sheet_table(ws, data, start_row=8)
        
        # Tạo footer với thông tin nguồn dữ liệu
        self._create_footer(ws, current_row + 2)
        
        # Điều chỉnh độ rộng cột
        self._adjust_column_width(ws)
        
        # Lưu file
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bang_can_doi_ke_toan_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_directory, filename)
        wb.save(filepath)
        
        print(f"✓ Đã tạo file bảng cân đối kế toán: {filepath}")
        return filepath
    
    def _create_header(self, ws, company_info):
        """Tạo header cho bảng cân đối kế toán"""
        
        # Tiêu đề chính
        ws.merge_cells('A1:D1')
        self._apply_cell_style(ws['A1'], 'title', 'BẢNG CÂN ĐỐI KẾ TOÁN')
        
        # Thông tin công ty
        ws.merge_cells('A2:D2')
        self._apply_cell_style(ws['A2'], 'subtitle', company_info['name'])
        
        # Kỳ báo cáo
        ws.merge_cells('A3:D3')
        self._apply_cell_style(ws['A3'], 'content', f"Tại ngày: {company_info['period']}")
        
        # Đơn vị tính
        ws.merge_cells('A4:D4')
        self._apply_cell_style(ws['A4'], 'content', f"Đơn vị tính: {company_info['unit']}")
        
        # Dòng trống
        ws.row_dimensions[5].height = 10
        
        # Header bảng
        headers = ['Chỉ tiêu', 'Mã số', 'Thuyết minh', 'Số cuối kỳ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            self._apply_cell_style(cell, 'header', header)
    
    def _create_balance_sheet_table(self, ws, data, start_row=7):
        """Tạo bảng cân đối kế toán chính"""
        current_row = start_row
        
        # PHẦN I: TÀI SẢN
        current_row = self._add_section_header(ws, current_row, "TÀI SẢN")
        
        # A. Tài sản ngắn hạn
        current_row = self._add_asset_section(ws, current_row, 
                                            data['assets']['A_TAI_SAN_NGAN_HAN'], 
                                            "100")
        
        # B. Tài sản dài hạn
        current_row = self._add_asset_section(ws, current_row, 
                                            data['assets']['B_TAI_SAN_DAI_HAN'], 
                                            "200")
        
        # Tổng cộng tài sản
        total_assets = self._calculate_total_assets(data['assets'])
        current_row = self._add_total_row(ws, current_row, "TỔNG CỘNG TÀI SẢN", "270", total_assets)
        
        # Dòng trống
        current_row += 1
        
        # PHẦN II: NGUỒN VỐN
        current_row = self._add_section_header(ws, current_row, "NGUỒN VỐN")
        
        # C. Nợ phải trả
        current_row = self._add_liability_section(ws, current_row, 
                                                data['liabilities_equity']['C_NO_PHAI_TRA'], 
                                                "300")
        
        # D. Vốn chủ sở hữu
        current_row = self._add_equity_section(ws, current_row, 
                                             data['liabilities_equity']['D_VON_CHU_SO_HUU'], 
                                             "400")
        
        # Tổng cộng nguồn vốn
        total_equity_liability = self._calculate_total_equity_liability(data['liabilities_equity'])
        current_row = self._add_total_row(ws, current_row, "TỔNG CỘNG NGUỒN VỐN", "440", total_equity_liability)
        
        return current_row
    
    def _add_section_header(self, ws, row, title):
        """Thêm header cho từng phần"""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1)
        self._apply_cell_style(cell, 'subtitle', title)
        return row + 1
    
    def _add_asset_section(self, ws, start_row, section_data, base_code):
        """Thêm section tài sản"""
        current_row = start_row
        
        # Header section
        ws.cell(row=current_row, column=1).value = section_data['description']
        self._apply_cell_style(ws.cell(row=current_row, column=1), 'content')
        self._apply_cell_style(ws.cell(row=current_row, column=2), 'content', base_code)
        self._apply_cell_style(ws.cell(row=current_row, column=3), 'content')
        
        # Tính tổng section
        section_total = sum(item['value'] for item in section_data['items'].values())
        self._apply_cell_style(ws.cell(row=current_row, column=4), 'total', section_total)
        
        current_row += 1
        
        # Chi tiết các khoản mục
        for code, item in section_data['items'].items():
            ws.cell(row=current_row, column=1).value = f"  - {item['name']}"
            self._apply_cell_style(ws.cell(row=current_row, column=1), 'content')
            self._apply_cell_style(ws.cell(row=current_row, column=2), 'content', code)
            self._apply_cell_style(ws.cell(row=current_row, column=3), 'content')
            self._apply_cell_style(ws.cell(row=current_row, column=4), 'number', item['value'])
            current_row += 1
        
        return current_row
    
    def _add_liability_section(self, ws, start_row, section_data, base_code):
        """Thêm section nợ phải trả"""
        return self._add_asset_section(ws, start_row, section_data, base_code)
    
    def _add_equity_section(self, ws, start_row, section_data, base_code):
        """Thêm section vốn chủ sở hữu"""
        return self._add_asset_section(ws, start_row, section_data, base_code)
    
    def _add_total_row(self, ws, row, title, code, value):
        """Thêm dòng tổng cộng"""
        self._apply_cell_style(ws.cell(row=row, column=1), 'content', title)
        self._apply_cell_style(ws.cell(row=row, column=2), 'content', code)
        self._apply_cell_style(ws.cell(row=row, column=3), 'content')
        self._apply_cell_style(ws.cell(row=row, column=4), 'total', value)
        return row + 1
    
    def _calculate_total_assets(self, assets_data):
        """Tính tổng tài sản"""
        total = 0
        for section in assets_data.values():
            total += sum(item['value'] for item in section['items'].values())
        return total
    
    def _calculate_total_equity_liability(self, equity_liability_data):
        """Tính tổng nguồn vốn"""
        total = 0
        for section in equity_liability_data.values():
            total += sum(item['value'] for item in section['items'].values())
        return total
    
    def _create_footer(self, ws, start_row):
        """Tạo footer với thông tin nguồn dữ liệu"""
        data_info = self.data_source.get_data_sources_info()
        
        ws.cell(row=start_row, column=1).value = "Nguồn dữ liệu:"
        self._apply_cell_style(ws.cell(row=start_row, column=1), 'content')
        
        ws.cell(row=start_row + 1, column=1).value = data_info['primary_source']
        self._apply_cell_style(ws.cell(row=start_row + 1, column=1), 'content')
        
        ws.cell(row=start_row + 2, column=1).value = f"Ngày tạo: {data_info['last_updated']}"
        self._apply_cell_style(ws.cell(row=start_row + 2, column=1), 'content')
        
        ws.cell(row=start_row + 3, column=1).value = data_info['disclaimer']
        self._apply_cell_style(ws.cell(row=start_row + 3, column=1), 'content')
    
    def _adjust_column_width(self, ws):
        """Điều chỉnh độ rộng cột"""
        column_widths = {
            'A': 40,  # Chỉ tiêu
            'B': 12,  # Mã số
            'C': 15,  # Thuyết minh
            'D': 20   # Số cuối kỳ
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def create_named_ranges(self, wb, filename):
        """
        Tạo named ranges để dễ dàng tham chiếu từ file khác
        Create named ranges for easy reference from other files
        """
        ws = wb.active
        
        # Tìm vị trí các tổng quan trọng trong bảng
        named_ranges = {
            'TotalAssets': None,
            'TotalLiabilities': None,
            'TotalEquity': None,
            'CurrentAssets': None,
            'CurrentLiabilities': None
        }
        
        # Quét qua các cell để tìm vị trí tổng
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value).upper()
                    if 'TỔNG CỘNG TÀI SẢN' in cell_value:
                        # Tổng tài sản ở cột D của dòng này
                        named_ranges['TotalAssets'] = f"D{cell.row}"
                    elif 'TÀI SẢN NGẮN HẠN' in cell_value and 'A.' in cell_value:
                        # Tài sản ngắn hạn
                        named_ranges['CurrentAssets'] = f"D{cell.row}"
        
        # Định nghĩa named ranges trong workbook
        for name, cell_ref in named_ranges.items():
            if cell_ref:
                wb.define_name(name, f"'{ws.title}'!{cell_ref}")
        
        return named_ranges

# Hàm tiện ích
def create_balance_sheet_file(output_dir="output", filename=None):
    """Hàm tiện ích để tạo file bảng cân đối kế toán"""
    generator = BalanceSheetGenerator(output_dir)
    return generator.create_balance_sheet(filename)

if __name__ == "__main__":
    # Test chức năng
    print("=== ĐANG TẠO FILE BẢNG CÂN ĐỐI KẾ TOÁN ===")
    
    generator = BalanceSheetGenerator("test_output")
    filepath = generator.create_balance_sheet("test_balance_sheet.xlsx")
    
    print(f"✓ Đã tạo thành công file: {filepath}")
    print("✓ File chứa:")
    print("  - Bảng cân đối kế toán chuẩn theo quy định Việt Nam")
    print("  - Dữ liệu mẫu với cấu trúc đầy đủ")
    print("  - Định dạng chuyên nghiệp với style phù hợp")
    print("  - Thông tin nguồn dữ liệu và disclaimer")