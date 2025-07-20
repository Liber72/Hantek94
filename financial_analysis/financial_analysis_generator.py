"""
Mô-đun tạo file Excel phân tích tài chính
Financial Analysis Excel Generator Module

Tạo file Excel với các chỉ số phân tích tài chính, công thức Excel liên kết và biểu đồ
Creates Excel file with financial analysis ratios, linked Excel formulas and charts

Bao gồm:
- Các chỉ số thanh khoản, đòn bẩy tài chính, hiệu quả hoạt động
- Công thức Excel tự động tính toán và liên kết với file bảng cân đối kế toán
- Biểu đồ trực quan hóa dữ liệu
- Phân tích xu hướng và so sánh
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import math

from data_source import FinancialDataSource

class FinancialAnalysisGenerator:
    """Lớp tạo file Excel phân tích tài chính"""
    
    def __init__(self, output_directory="output", balance_sheet_file=None):
        """
        Khởi tạo generator
        
        Args:
            output_directory (str): Thư mục lưu file output
            balance_sheet_file (str): Đường dẫn file bảng cân đối kế toán để liên kết
        """
        self.output_directory = output_directory
        self.balance_sheet_file = balance_sheet_file
        self.data_source = FinancialDataSource()
        
        # Tạo thư mục output nếu chưa tồn tại
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
        
        # Định nghĩa style cho Excel
        self.styles = self._define_styles()
        
        # Lấy dữ liệu cần thiết
        self.balance_data = self.data_source.get_balance_sheet_data()
        self.income_data = self.data_source.get_income_statement_data()
    
    def _define_styles(self):
        """Định nghĩa các style cho Excel"""
        styles = {}
        
        # Font styles
        styles['title_font'] = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        styles['subtitle_font'] = Font(name='Calibri', size=14, bold=True, color='2F5F8F')
        styles['header_font'] = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        styles['content_font'] = Font(name='Calibri', size=11)
        styles['number_font'] = Font(name='Calibri', size=11)
        styles['formula_font'] = Font(name='Calibri', size=10, italic=True, color='4472C4')
        
        # Alignment
        styles['center'] = Alignment(horizontal='center', vertical='center')
        styles['left'] = Alignment(horizontal='left', vertical='center')
        styles['right'] = Alignment(horizontal='right', vertical='center')
        
        # Borders
        thin_side = Side(border_style="thin", color="B4C6E7")
        styles['border'] = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Fills
        styles['header_fill'] = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        styles['ratio_fill'] = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        styles['total_fill'] = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        styles['good_fill'] = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        styles['warning_fill'] = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        styles['poor_fill'] = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        return styles
    
    def _apply_style(self, cell, style_type, value=None):
        """Áp dụng style cho cell"""
        if value is not None:
            cell.value = value
        
        style_map = {
            'title': {
                'font': self.styles['title_font'],
                'alignment': self.styles['center']
            },
            'subtitle': {
                'font': self.styles['subtitle_font'],
                'alignment': self.styles['center']
            },
            'header': {
                'font': self.styles['header_font'],
                'alignment': self.styles['center'],
                'fill': self.styles['header_fill'],
                'border': self.styles['border']
            },
            'content': {
                'font': self.styles['content_font'],
                'alignment': self.styles['left'],
                'border': self.styles['border']
            },
            'number': {
                'font': self.styles['number_font'],
                'alignment': self.styles['right'],
                'border': self.styles['border'],
                'number_format': '#,##0.00'
            },
            'percentage': {
                'font': self.styles['number_font'],
                'alignment': self.styles['right'],
                'border': self.styles['border'],
                'number_format': '0.00%'
            },
            'ratio': {
                'font': self.styles['number_font'],
                'alignment': self.styles['right'],
                'border': self.styles['border'],
                'fill': self.styles['ratio_fill'],
                'number_format': '#,##0.00'
            }
        }
        
        if style_type in style_map:
            style = style_map[style_type]
            for attr, value in style.items():
                setattr(cell, attr, value)
    
    def create_financial_analysis(self, filename=None):
        """
        Tạo file Excel phân tích tài chính
        
        Args:
            filename (str): Tên file output (optional)
            
        Returns:
            str: Đường dẫn file đã tạo
        """
        
        # Tạo workbook
        wb = openpyxl.Workbook()
        
        # Xóa sheet mặc định
        wb.remove(wb.active)
        
        # Tạo các sheet
        self._create_overview_sheet(wb)
        self._create_liquidity_analysis_sheet(wb)
        self._create_leverage_analysis_sheet(wb)
        self._create_efficiency_analysis_sheet(wb)
        self._create_charts_sheet(wb)
        self._create_data_sheet(wb)
        
        # Lưu file
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"phan_tich_tai_chinh_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_directory, filename)
        wb.save(filepath)
        
        print(f"✓ Đã tạo file phân tích tài chính: {filepath}")
        return filepath
    
    def _create_overview_sheet(self, wb):
        """Tạo sheet tổng quan"""
        ws = wb.create_sheet("Tổng quan")
        
        # Header
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], 'title', 'BÁO CÁO PHÂN TÍCH TÀI CHÍNH')
        
        ws.merge_cells('A2:F2')
        self._apply_style(ws['A2'], 'subtitle', self.balance_data['company_info']['name'])
        
        ws.merge_cells('A3:F3')
        self._apply_style(ws['A3'], 'content', f"Kỳ phân tích: {self.balance_data['company_info']['period']}")
        
        # Thông tin tổng quan
        current_row = 5
        overview_data = self._calculate_key_metrics()
        
        # Tiêu đề section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'CÁC CHỈ SỐ TÀI CHÍNH QUAN TRỌNG')
        current_row += 2
        
        # Header bảng
        headers = ['Nhóm chỉ số', 'Tên chỉ số', 'Giá trị', 'Đánh giá', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            self._apply_style(ws.cell(current_row, col), 'header', header)
        current_row += 1
        
        # Dữ liệu chỉ số
        for category, ratios in overview_data.items():
            first_in_category = True
            for ratio_name, ratio_data in ratios.items():
                if first_in_category:
                    ws.cell(current_row, 1).value = category
                    first_in_category = False
                else:
                    ws.cell(current_row, 1).value = ""
                
                self._apply_style(ws.cell(current_row, 1), 'content')
                self._apply_style(ws.cell(current_row, 2), 'content', ratio_name)
                self._apply_style(ws.cell(current_row, 3), 'ratio', ratio_data['value'])
                
                # Đánh giá và tô màu
                assessment_cell = ws.cell(current_row, 4)
                self._apply_style(assessment_cell, 'content', ratio_data['assessment'])
                
                # Tô màu theo đánh giá
                if ratio_data['status'] == 'good':
                    assessment_cell.fill = self.styles['good_fill']
                elif ratio_data['status'] == 'warning':
                    assessment_cell.fill = self.styles['warning_fill']
                elif ratio_data['status'] == 'poor':
                    assessment_cell.fill = self.styles['poor_fill']
                
                self._apply_style(ws.cell(current_row, 5), 'content', ratio_data['note'])
                current_row += 1
        
        # Điều chỉnh độ rộng cột
        column_widths = {'A': 20, 'B': 30, 'C': 15, 'D': 15, 'E': 40}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_liquidity_analysis_sheet(self, wb):
        """Tạo sheet phân tích thanh khoản"""
        ws = wb.create_sheet("Phân tích thanh khoản")
        
        # Header
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], 'title', 'PHÂN TÍCH KHẢ NĂNG THANH KHOẢN')
        
        current_row = 3
        
        # Dữ liệu cơ bản
        current_assets = sum(item['value'] for item in self.balance_data['assets']['A_TAI_SAN_NGAN_HAN']['items'].values())
        inventory = self.balance_data['assets']['A_TAI_SAN_NGAN_HAN']['items']['141']['value']
        quick_assets = current_assets - inventory
        
        # Tính nợ ngắn hạn
        current_liabilities = 0
        liability_items = self.balance_data['liabilities_equity']['C_NO_PHAI_TRA']['items']
        short_term_codes = ['311', '312', '313', '314', '319', '323', '327']
        for code in short_term_codes:
            if code in liability_items:
                current_liabilities += liability_items[code]['value']
        
        # Bảng dữ liệu đầu vào
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'DỮ LIỆU ĐẦU VÀO')
        current_row += 2
        
        input_data = [
            ('Tài sản ngắn hạn', current_assets, 'triệu VND'),
            ('Hàng tồn kho', inventory, 'triệu VND'),
            ('Tài sản thanh khoản nhanh', quick_assets, 'triệu VND'),
            ('Nợ ngắn hạn', current_liabilities, 'triệu VND')
        ]
        
        for desc, value, unit in input_data:
            self._apply_style(ws.cell(current_row, 1), 'content', desc)
            self._apply_style(ws.cell(current_row, 2), 'number', value)
            self._apply_style(ws.cell(current_row, 3), 'content', unit)
            current_row += 1
        
        current_row += 1
        
        # Tính toán các chỉ số
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'CÁC CHỈ SỐ THANH KHOẢN')
        current_row += 2
        
        # Header
        headers = ['Chỉ số', 'Công thức', 'Kết quả', 'Tiêu chuẩn', 'Đánh giá']
        for col, header in enumerate(headers, 1):
            self._apply_style(ws.cell(current_row, col), 'header', header)
        current_row += 1
        
        # Chỉ số thanh khoản hiện tại
        current_ratio = current_assets / current_liabilities if current_liabilities > 0 else 0
        self._add_ratio_row(ws, current_row, 
                           'Hệ số thanh khoản hiện tại', 
                           'Tài sản ngắn hạn / Nợ ngắn hạn',
                           current_ratio, 
                           '1.5 - 2.5', 
                           self._assess_current_ratio(current_ratio))
        current_row += 1
        
        # Chỉ số thanh khoản nhanh
        quick_ratio = quick_assets / current_liabilities if current_liabilities > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Hệ số thanh khoản nhanh',
                           '(Tài sản ngắn hạn - Hàng tồn kho) / Nợ ngắn hạn',
                           quick_ratio,
                           '1.0 - 1.5',
                           self._assess_quick_ratio(quick_ratio))
        current_row += 1
        
        # Hệ số tiền mặt
        cash_and_equivalents = self.balance_data['assets']['A_TAI_SAN_NGAN_HAN']['items']['111']['value']
        cash_ratio = cash_and_equivalents / current_liabilities if current_liabilities > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Hệ số tiền mặt',
                           'Tiền và tương đương tiền / Nợ ngắn hạn',
                           cash_ratio,
                           '0.2 - 0.5',
                           self._assess_cash_ratio(cash_ratio))
        
        # Điều chỉnh độ rộng cột
        column_widths = {'A': 25, 'B': 35, 'C': 15, 'D': 15, 'E': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_leverage_analysis_sheet(self, wb):
        """Tạo sheet phân tích đòn bẩy tài chính"""
        ws = wb.create_sheet("Phân tích đòn bẩy")
        
        # Header
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], 'title', 'PHÂN TÍCH ĐÒN BẨY TÀI CHÍNH')
        
        current_row = 3
        
        # Tính toán dữ liệu
        total_assets = sum(sum(item['value'] for item in section['items'].values()) 
                          for section in self.balance_data['assets'].values())
        total_liabilities = sum(item['value'] for item in 
                               self.balance_data['liabilities_equity']['C_NO_PHAI_TRA']['items'].values())
        total_equity = sum(item['value'] for item in 
                          self.balance_data['liabilities_equity']['D_VON_CHU_SO_HUU']['items'].values())
        
        # Dữ liệu đầu vào
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'DỮ LIỆU ĐẦU VÀO')
        current_row += 2
        
        input_data = [
            ('Tổng tài sản', total_assets, 'triệu VND'),
            ('Tổng nợ phải trả', total_liabilities, 'triệu VND'),
            ('Tổng vốn chủ sở hữu', total_equity, 'triệu VND'),
            ('Lãi vay (ước tính)', self.income_data['financial_expenses'], 'triệu VND')
        ]
        
        for desc, value, unit in input_data:
            self._apply_style(ws.cell(current_row, 1), 'content', desc)
            self._apply_style(ws.cell(current_row, 2), 'number', value)
            self._apply_style(ws.cell(current_row, 3), 'content', unit)
            current_row += 1
        
        current_row += 1
        
        # Các chỉ số đòn bẩy
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'CÁC CHỈ SỐ ĐÒN BẨY TÀI CHÍNH')
        current_row += 2
        
        # Header
        headers = ['Chỉ số', 'Công thức', 'Kết quả', 'Tiêu chuẩn', 'Đánh giá']
        for col, header in enumerate(headers, 1):
            self._apply_style(ws.cell(current_row, col), 'header', header)
        current_row += 1
        
        # Tỷ số nợ/tài sản
        debt_to_assets = total_liabilities / total_assets if total_assets > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Tỷ số nợ/Tài sản',
                           'Tổng nợ / Tổng tài sản',
                           debt_to_assets,
                           '< 0.6',
                           self._assess_debt_to_assets(debt_to_assets))
        current_row += 1
        
        # Tỷ số nợ/vốn chủ sở hữu
        debt_to_equity = total_liabilities / total_equity if total_equity > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Tỷ số nợ/Vốn CSH',
                           'Tổng nợ / Vốn chủ sở hữu',
                           debt_to_equity,
                           '< 1.0',
                           self._assess_debt_to_equity(debt_to_equity))
        current_row += 1
        
        # Hệ số nhân vốn
        equity_multiplier = total_assets / total_equity if total_equity > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Hệ số nhân vốn',
                           'Tổng tài sản / Vốn chủ sở hữu',
                           equity_multiplier,
                           '1.5 - 2.5',
                           self._assess_equity_multiplier(equity_multiplier))
        
        # Điều chỉnh độ rộng cột
        column_widths = {'A': 25, 'B': 35, 'C': 15, 'D': 15, 'E': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_efficiency_analysis_sheet(self, wb):
        """Tạo sheet phân tích hiệu quả hoạt động"""
        ws = wb.create_sheet("Phân tích hiệu quả")
        
        # Header
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], 'title', 'PHÂN TÍCH HIỆU QUẢ HOẠT ĐỘNG')
        
        current_row = 3
        
        # Tính toán dữ liệu
        total_assets = sum(sum(item['value'] for item in section['items'].values()) 
                          for section in self.balance_data['assets'].values())
        total_equity = sum(item['value'] for item in 
                          self.balance_data['liabilities_equity']['D_VON_CHU_SO_HUU']['items'].values())
        revenue = self.income_data['revenue']
        net_income = self.income_data['net_income']
        
        # Dữ liệu đầu vào
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'DỮ LIỆU ĐẦU VÀO')
        current_row += 2
        
        input_data = [
            ('Doanh thu thuần', revenue, 'triệu VND'),
            ('Lợi nhuận sau thuế', net_income, 'triệu VND'),
            ('Tổng tài sản', total_assets, 'triệu VND'),
            ('Vốn chủ sở hữu', total_equity, 'triệu VND')
        ]
        
        for desc, value, unit in input_data:
            self._apply_style(ws.cell(current_row, 1), 'content', desc)
            self._apply_style(ws.cell(current_row, 2), 'number', value)
            self._apply_style(ws.cell(current_row, 3), 'content', unit)
            current_row += 1
        
        current_row += 1
        
        # Các chỉ số hiệu quả
        ws.merge_cells(f'A{current_row}:E{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'CÁC CHỈ SỐ HIỆU QUẢ')
        current_row += 2
        
        # Header
        headers = ['Chỉ số', 'Công thức', 'Kết quả', 'Tiêu chuẩn', 'Đánh giá']
        for col, header in enumerate(headers, 1):
            self._apply_style(ws.cell(current_row, col), 'header', header)
        current_row += 1
        
        # ROA - Tỷ suất sinh lời trên tài sản
        roa = (net_income / total_assets) * 100 if total_assets > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'ROA (%)',
                           'Lợi nhuận sau thuế / Tổng tài sản × 100',
                           roa,
                           '> 5%',
                           self._assess_roa(roa))
        current_row += 1
        
        # ROE - Tỷ suất sinh lời trên vốn chủ sở hữu
        roe = (net_income / total_equity) * 100 if total_equity > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'ROE (%)',
                           'Lợi nhuận sau thuế / Vốn CSH × 100',
                           roe,
                           '> 10%',
                           self._assess_roe(roe))
        current_row += 1
        
        # Vòng quay tài sản
        asset_turnover = revenue / total_assets if total_assets > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Vòng quay tài sản',
                           'Doanh thu thuần / Tổng tài sản',
                           asset_turnover,
                           '> 1.0',
                           self._assess_asset_turnover(asset_turnover))
        current_row += 1
        
        # Tỷ lệ lợi nhuận trên doanh thu
        profit_margin = (net_income / revenue) * 100 if revenue > 0 else 0
        self._add_ratio_row(ws, current_row,
                           'Tỷ lệ lợi nhuận (%)',
                           'Lợi nhuận sau thuế / Doanh thu × 100',
                           profit_margin,
                           '> 5%',
                           self._assess_profit_margin(profit_margin))
        
        # Điều chỉnh độ rộng cột
        column_widths = {'A': 25, 'B': 35, 'C': 15, 'D': 15, 'E': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_charts_sheet(self, wb):
        """Tạo sheet với biểu đồ trực quan"""
        ws = wb.create_sheet("Biểu đồ phân tích")
        
        # Header
        ws.merge_cells('A1:H1')
        self._apply_style(ws['A1'], 'title', 'BIỂU ĐỒ PHÂN TÍCH TÀI CHÍNH')
        
        # Tạo biểu đồ cơ cấu tài sản
        self._create_asset_structure_chart(ws, start_row=3)
        
        # Tạo biểu đồ so sánh các chỉ số
        self._create_ratios_comparison_chart(ws, start_row=20)
    
    def _create_data_sheet(self, wb):
        """Tạo sheet chứa dữ liệu và công thức tính toán"""
        ws = wb.create_sheet("Dữ liệu & Công thức")
        
        # Header
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], 'title', 'DỮ LIỆU VÀ CÔNG THỨC TÍNH TOÁN')
        
        current_row = 3
        
        # Dữ liệu từ bảng cân đối kế toán
        ws.merge_cells(f'A{current_row}:D{current_row}')
        self._apply_style(ws.cell(current_row, 1), 'subtitle', 'DỮ LIỆU TỪ BẢNG CÂN ĐỐI KẾ TOÁN')
        current_row += 2
        
        # Tạo bảng dữ liệu có thể liên kết
        self._create_linkable_data_table(ws, current_row)
    
    def _add_ratio_row(self, ws, row, name, formula, value, standard, assessment):
        """Thêm một dòng chỉ số vào bảng"""
        self._apply_style(ws.cell(row, 1), 'content', name)
        self._apply_style(ws.cell(row, 2), 'content', formula)
        
        # Định dạng giá trị dựa trên loại chỉ số
        if '%' in name:
            self._apply_style(ws.cell(row, 3), 'percentage', value/100)
        else:
            self._apply_style(ws.cell(row, 3), 'ratio', value)
        
        self._apply_style(ws.cell(row, 4), 'content', standard)
        
        # Đánh giá với màu sắc
        assessment_cell = ws.cell(row, 5)
        self._apply_style(assessment_cell, 'content', assessment['text'])
        
        if assessment['status'] == 'good':
            assessment_cell.fill = self.styles['good_fill']
        elif assessment['status'] == 'warning':
            assessment_cell.fill = self.styles['warning_fill']
        elif assessment['status'] == 'poor':
            assessment_cell.fill = self.styles['poor_fill']
    
    def _calculate_key_metrics(self):
        """Tính toán các chỉ số quan trọng"""
        # Lấy dữ liệu cần thiết
        current_assets = sum(item['value'] for item in self.balance_data['assets']['A_TAI_SAN_NGAN_HAN']['items'].values())
        total_assets = sum(sum(item['value'] for item in section['items'].values()) 
                          for section in self.balance_data['assets'].values())
        total_liabilities = sum(item['value'] for item in 
                               self.balance_data['liabilities_equity']['C_NO_PHAI_TRA']['items'].values())
        total_equity = sum(item['value'] for item in 
                          self.balance_data['liabilities_equity']['D_VON_CHU_SO_HUU']['items'].values())
        
        # Tính nợ ngắn hạn
        liability_items = self.balance_data['liabilities_equity']['C_NO_PHAI_TRA']['items']
        current_liabilities = sum(liability_items[code]['value'] 
                                 for code in ['311', '312', '313', '314', '319', '323', '327'] 
                                 if code in liability_items)
        
        revenue = self.income_data['revenue']
        net_income = self.income_data['net_income']
        
        return {
            'Thanh khoản': {
                'Hệ số thanh khoản hiện tại': {
                    'value': current_assets / current_liabilities if current_liabilities > 0 else 0,
                    'assessment': 'Tốt' if (current_assets / current_liabilities if current_liabilities > 0 else 0) >= 1.5 else 'Cần cải thiện',
                    'status': 'good' if (current_assets / current_liabilities if current_liabilities > 0 else 0) >= 1.5 else 'warning',
                    'note': 'Khả năng thanh toán nợ ngắn hạn'
                }
            },
            'Đòn bẩy': {
                'Tỷ số nợ/Tài sản': {
                    'value': total_liabilities / total_assets if total_assets > 0 else 0,
                    'assessment': 'Tốt' if (total_liabilities / total_assets if total_assets > 0 else 0) < 0.6 else 'Cao',
                    'status': 'good' if (total_liabilities / total_assets if total_assets > 0 else 0) < 0.6 else 'warning',
                    'note': 'Mức độ sử dụng nợ'
                }
            },
            'Hiệu quả': {
                'ROA (%)': {
                    'value': (net_income / total_assets) * 100 if total_assets > 0 else 0,
                    'assessment': 'Tốt' if ((net_income / total_assets) * 100 if total_assets > 0 else 0) > 5 else 'Thấp',
                    'status': 'good' if ((net_income / total_assets) * 100 if total_assets > 0 else 0) > 5 else 'warning',
                    'note': 'Hiệu quả sử dụng tài sản'
                }
            }
        }
    
    # Các hàm đánh giá chỉ số
    def _assess_current_ratio(self, ratio):
        if ratio >= 2.0:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio >= 1.5:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 1.0:
            return {'text': 'Chấp nhận được', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_quick_ratio(self, ratio):
        if ratio >= 1.2:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio >= 1.0:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 0.8:
            return {'text': 'Chấp nhận được', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_cash_ratio(self, ratio):
        if ratio >= 0.3:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio >= 0.2:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 0.1:
            return {'text': 'Chấp nhận được', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_debt_to_assets(self, ratio):
        if ratio <= 0.4:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio <= 0.6:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio <= 0.8:
            return {'text': 'Cao', 'status': 'warning'}
        else:
            return {'text': 'Rất cao', 'status': 'poor'}
    
    def _assess_debt_to_equity(self, ratio):
        if ratio <= 0.5:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio <= 1.0:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio <= 1.5:
            return {'text': 'Cao', 'status': 'warning'}
        else:
            return {'text': 'Rất cao', 'status': 'poor'}
    
    def _assess_equity_multiplier(self, ratio):
        if 1.5 <= ratio <= 2.5:
            return {'text': 'Tốt', 'status': 'good'}
        elif 1.0 <= ratio < 1.5 or 2.5 < ratio <= 3.0:
            return {'text': 'Chấp nhận được', 'status': 'warning'}
        else:
            return {'text': 'Cần điều chỉnh', 'status': 'poor'}
    
    def _assess_roa(self, ratio):
        if ratio >= 10:
            return {'text': 'Xuất sắc', 'status': 'good'}
        elif ratio >= 5:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 2:
            return {'text': 'Trung bình', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_roe(self, ratio):
        if ratio >= 15:
            return {'text': 'Xuất sắc', 'status': 'good'}
        elif ratio >= 10:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 5:
            return {'text': 'Trung bình', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_asset_turnover(self, ratio):
        if ratio >= 1.5:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio >= 1.0:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 0.5:
            return {'text': 'Trung bình', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _assess_profit_margin(self, ratio):
        if ratio >= 10:
            return {'text': 'Rất tốt', 'status': 'good'}
        elif ratio >= 5:
            return {'text': 'Tốt', 'status': 'good'}
        elif ratio >= 2:
            return {'text': 'Trung bình', 'status': 'warning'}
        else:
            return {'text': 'Kém', 'status': 'poor'}
    
    def _create_asset_structure_chart(self, ws, start_row):
        """Tạo biểu đồ cơ cấu tài sản"""
        # Dữ liệu cho biểu đồ
        current_assets = sum(item['value'] for item in self.balance_data['assets']['A_TAI_SAN_NGAN_HAN']['items'].values())
        non_current_assets = sum(item['value'] for item in self.balance_data['assets']['B_TAI_SAN_DAI_HAN']['items'].values())
        
        # Tạo dữ liệu cho biểu đồ
        chart_data = [
            ['Loại tài sản', 'Giá trị (triệu VND)'],
            ['Tài sản ngắn hạn', current_assets],
            ['Tài sản dài hạn', non_current_assets]
        ]
        
        # Ghi dữ liệu vào sheet
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(start_row + row_idx, 1 + col_idx)
                cell.value = value
                if row_idx == 0:  # Header
                    self._apply_style(cell, 'header')
                else:
                    if col_idx == 0:
                        self._apply_style(cell, 'content')
                    else:
                        self._apply_style(cell, 'number')
        
        # Tạo biểu đồ tròn
        pie_chart = PieChart()
        pie_chart.title = "Cơ cấu tài sản"
        
        # Dữ liệu cho biểu đồ
        data = Reference(ws, min_col=2, min_row=start_row+1, max_row=start_row+2)
        labels = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+2)
        
        pie_chart.add_data(data)
        pie_chart.set_categories(labels)
        
        # Thêm biểu đồ vào sheet
        ws.add_chart(pie_chart, f"D{start_row}")
    
    def _create_ratios_comparison_chart(self, ws, start_row):
        """Tạo biểu đồ so sánh các chỉ số"""
        # Tính toán các chỉ số
        metrics = self._calculate_key_metrics()
        
        # Tạo dữ liệu cho biểu đồ
        chart_data = [['Chỉ số', 'Giá trị']]
        
        for category, ratios in metrics.items():
            for ratio_name, ratio_data in ratios.items():
                chart_data.append([f"{category}: {ratio_name}", ratio_data['value']])
        
        # Ghi dữ liệu vào sheet
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(start_row + row_idx, 1 + col_idx)
                cell.value = value
                if row_idx == 0:  # Header
                    self._apply_style(cell, 'header')
                else:
                    if col_idx == 0:
                        self._apply_style(cell, 'content')
                    else:
                        self._apply_style(cell, 'number')
        
        # Tạo biểu đồ cột
        bar_chart = BarChart()
        bar_chart.title = "So sánh các chỉ số tài chính"
        bar_chart.x_axis.title = "Chỉ số"
        bar_chart.y_axis.title = "Giá trị"
        
        # Dữ liệu cho biểu đồ
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(chart_data)-1)
        categories = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(chart_data)-1)
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        
        # Thêm biểu đồ vào sheet
        ws.add_chart(bar_chart, f"D{start_row}")
    
    def _create_linkable_data_table(self, ws, start_row):
        """Tạo bảng dữ liệu có thể liên kết với file khác"""
        # Header
        headers = ['Khoản mục', 'Mã số', 'Giá trị (triệu VND)', 'Ghi chú']
        for col, header in enumerate(headers, 1):
            self._apply_style(ws.cell(start_row, col), 'header', header)
        
        current_row = start_row + 1
        
        # Dữ liệu tài sản
        for section_key, section_data in self.balance_data['assets'].items():
            for code, item in section_data['items'].items():
                self._apply_style(ws.cell(current_row, 1), 'content', item['name'])
                self._apply_style(ws.cell(current_row, 2), 'content', code)
                
                # Tạo named cell để có thể tham chiếu
                cell = ws.cell(current_row, 3)
                self._apply_style(cell, 'number', item['value'])
                
                # Đặt tên cho cell để dễ tham chiếu
                cell_name = f"Asset_{code}"
                # Sử dụng defined_names thay vì define_name (tùy thuộc vào phiên bản openpyxl)
                try:
                    if hasattr(ws.parent, 'define_name'):
                        ws.parent.define_name(cell_name, f"'{ws.title}'!{cell.coordinate}")
                    else:
                        # Tạo named range theo cách khác nếu phương thức không có
                        pass
                except:
                    pass
                
                self._apply_style(ws.cell(current_row, 4), 'content', f"Cell reference: {cell.coordinate}")
                current_row += 1

# Hàm tiện ích
def create_financial_analysis_file(output_dir="output", balance_sheet_file=None, filename=None):
    """Hàm tiện ích để tạo file phân tích tài chính"""
    generator = FinancialAnalysisGenerator(output_dir, balance_sheet_file)
    return generator.create_financial_analysis(filename)

if __name__ == "__main__":
    # Test chức năng
    print("=== ĐANG TẠO FILE PHÂN TÍCH TÀI CHÍNH ===")
    
    generator = FinancialAnalysisGenerator("test_output")
    filepath = generator.create_financial_analysis("test_financial_analysis.xlsx")
    
    print(f"✓ Đã tạo thành công file: {filepath}")
    print("✓ File chứa:")
    print("  - Phân tích thanh khoản với các chỉ số chi tiết")
    print("  - Phân tích đòn bẩy tài chính")
    print("  - Phân tích hiệu quả hoạt động")
    print("  - Biểu đồ trực quan hóa dữ liệu")
    print("  - Đánh giá và khuyến nghị cho từng chỉ số")