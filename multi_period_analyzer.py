"""
Multi-Period Analyzer - Hệ thống Phân tích Nhiều Kỳ
====================================================

Phân tích so sánh dữ liệu tài chính qua nhiều kỳ báo cáo
với công thức Excel động và tính toán xu hướng tăng trưởng

Tác giả: Hệ thống Phân tích Tài chính Động
Chuẩn: VAS/Circular 200/2014/TT-BTC
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
import datetime
import json
import os

class MultiPeriodAnalyzer:
    def __init__(self, periods=3):
        self.periods = periods  # Số kỳ so sánh (mặc định 3 kỳ)
        self.wb = None
        self.sheets = {}
        
        # Thiết lập style
        self.setup_styles()
        
        # Dữ liệu mẫu cho nhiều kỳ
        self.setup_multi_period_data()
        
    def setup_styles(self):
        """Thiết lập các style Excel chuyên nghiệp"""
        self.font_header = Font(name='Times New Roman', size=14, bold=True, color='FFFFFF')
        self.font_title = Font(name='Times New Roman', size=12, bold=True)
        self.font_normal = Font(name='Times New Roman', size=11)
        self.font_bold = Font(name='Times New Roman', size=11, bold=True)
        
        # Màu sắc theo từng kỳ
        self.fill_header = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        self.fill_current = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.fill_previous = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.fill_trend_up = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.fill_trend_down = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Border
        self.border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
    def setup_multi_period_data(self):
        """Thiết lập dữ liệu mẫu cho nhiều kỳ"""
        # Dữ liệu 3 kỳ: 2022, 2023, 2024
        self.period_data = {
            '2022': {
                'total_assets': 120000000000,
                'current_assets': 50000000000,
                'current_liabilities': 30000000000,
                'total_equity': 60000000000,
                'revenue': 70000000000,
                'net_income': 7000000000,
                'inventory': 15000000000,
                'accounts_receivable': 10000000000
            },
            '2023': {
                'total_assets': 130000000000,
                'current_assets': 58000000000,
                'current_liabilities': 35000000000,
                'total_equity': 65000000000,
                'revenue': 80000000000,
                'net_income': 8640000000,
                'inventory': 18000000000,
                'accounts_receivable': 12000000000
            },
            '2024': {
                'total_assets': 145000000000,
                'current_assets': 68000000000,
                'current_liabilities': 40000000000,
                'total_equity': 75000000000,
                'revenue': 92000000000,
                'net_income': 11040000000,
                'inventory': 20000000000,
                'accounts_receivable': 14000000000
            }
        }
        
        # Danh sách các kỳ
        self.periods_list = list(self.period_data.keys())
        
    def create_multi_period_workbook(self):
        """Tạo workbook phân tích nhiều kỳ"""
        print("🚀 Bắt đầu tạo hệ thống phân tích nhiều kỳ...")
        
        self.wb = openpyxl.Workbook()
        
        # Tạo các sheet
        self.create_multi_period_balance_sheet()
        self.create_trend_analysis_sheet()
        self.create_growth_analysis_sheet()
        self.create_ratio_comparison_sheet()
        self.create_forecast_sheet()
        
        print("✅ Hoàn thành tạo workbook phân tích nhiều kỳ")
        
    def create_multi_period_balance_sheet(self):
        """Tạo bảng cân đối kế toán nhiều kỳ"""
        print("📊 Đang tạo bảng cân đối nhiều kỳ...")
        
        ws = self.wb.active
        ws.title = "Bảng Cân Đối Nhiều Kỳ"
        self.sheets['balance'] = ws
        
        # Header
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "BẢNG CÂN ĐỐI KẾ TOÁN NHIỀU KỲ SO SÁNH"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Header cột
        row = 3
        ws[f'A{row}'] = "Chỉ tiêu"
        col = 2
        for period in self.periods_list:
            ws.cell(row=row, column=col).value = f"Năm {period}"
            ws.cell(row=row, column=col).font = self.font_bold
            ws.cell(row=row, column=col).fill = self.fill_current
            ws.cell(row=row, column=col).border = self.border_thin
            col += 1
            
        # Cột tăng trưởng
        ws.cell(row=row, column=col).value = "Tăng trưởng (%)"
        ws.cell(row=row, column=col).font = self.font_bold
        ws.cell(row=row, column=col).fill = self.fill_trend_up
        ws.cell(row=row, column=col).border = self.border_thin
        
        row += 1
        
        # Dữ liệu tài sản
        balance_items = [
            ('Tổng tài sản', 'total_assets'),
            ('Tài sản ngắn hạn', 'current_assets'),
            ('Nợ ngắn hạn', 'current_liabilities'),
            ('Vốn chủ sở hữu', 'total_equity'),
            ('Doanh thu', 'revenue'),
            ('Lợi nhuận sau thuế', 'net_income'),
            ('Hàng tồn kho', 'inventory'),
            ('Phải thu khách hàng', 'accounts_receivable')
        ]
        
        for item_name, data_key in balance_items:
            ws[f'A{row}'] = item_name
            ws[f'A{row}'].font = self.font_bold
            
            col = 2
            values = []
            for period in self.periods_list:
                value = self.period_data[period][data_key]
                values.append(value)
                ws.cell(row=row, column=col).value = value
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).border = self.border_thin
                
                # Tạo named range cho từng kỳ
                range_name = f"{data_key}_{period}"
                cell_ref = f"'{ws.title}'!{ws.cell(row=row, column=col).coordinate}"
                self.create_named_range(range_name, cell_ref)
                
                col += 1
                
            # Tính tăng trưởng (công thức Excel)
            if len(values) >= 2:
                first_col = chr(66)  # Cột B
                last_col = chr(66 + len(self.periods_list) - 1)
                growth_formula = f"=({last_col}{row}-{first_col}{row})/{first_col}{row}*100"
                ws.cell(row=row, column=col).value = growth_formula
                ws.cell(row=row, column=col).number_format = '0.00%'
                ws.cell(row=row, column=col).border = self.border_thin
                
            row += 1
            
        # Định dạng cột
        ws.column_dimensions['A'].width = 25
        for i in range(2, 2 + len(self.periods_list) + 1):
            ws.column_dimensions[chr(64 + i)].width = 18
            
        print("✅ Hoàn thành bảng cân đối nhiều kỳ")
        
    def create_trend_analysis_sheet(self):
        """Tạo sheet phân tích xu hướng"""
        print("📈 Đang tạo sheet Phân tích Xu hướng...")
        
        ws = self.wb.create_sheet("Phân Tích Xu Hướng")
        self.sheets['trend'] = ws
        
        # Header
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = "PHÂN TÍCH XU HƯỚNG PHÁT TRIỂN"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Bảng các chỉ số tài chính qua các kỳ
        row = 3
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = "CÁC CHỈ SỐ TÀI CHÍNH QUA CÁC KỲ"
        cell.font = self.font_title
        cell.fill = self.fill_current
        row += 1
        
        # Header
        headers = ['Chỉ số'] + [f'Năm {p}' for p in self.periods_list] + ['Xu hướng', 'Đánh giá']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_current
            cell.border = self.border_thin
        row += 1
        
        # Các chỉ số tài chính với công thức động
        financial_ratios = [
            ('Current Ratio', 'current_assets', 'current_liabilities', '={}/{}'),
            ('ROA (%)', 'net_income', 'total_assets', '={}*100/{}'),
            ('ROE (%)', 'net_income', 'total_equity', '={}*100/{}'),
            ('Asset Turnover', 'revenue', 'total_assets', '={}/{}'),
            ('Inventory Turnover', 'revenue', 'inventory', '={}/{}'),
            ('Revenue Growth (%)', 'revenue', None, None)
        ]
        
        for ratio_name, numerator, denominator, formula_template in financial_ratios:
            ws[f'A{row}'] = ratio_name
            ws[f'A{row}'].font = self.font_bold
            
            col = 2
            ratio_values = []
            
            for period in self.periods_list:
                if denominator:
                    # Tính tỷ số
                    num_val = self.period_data[period][numerator]
                    den_val = self.period_data[period][denominator]
                    
                    if ratio_name in ['ROA (%)', 'ROE (%)']:
                        ratio_val = (num_val / den_val) * 100
                    else:
                        ratio_val = num_val / den_val
                else:
                    # Đối với revenue growth, tính riêng
                    ratio_val = self.period_data[period][numerator] / 1000000000  # Tỷ VND
                    
                ratio_values.append(ratio_val)
                ws.cell(row=row, column=col).value = ratio_val
                
                if ratio_name in ['ROA (%)', 'ROE (%)']:
                    ws.cell(row=row, column=col).number_format = '0.00%'
                else:
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                    
                ws.cell(row=row, column=col).border = self.border_thin
                col += 1
                
            # Phân tích xu hướng
            if len(ratio_values) >= 2:
                trend = "Tăng" if ratio_values[-1] > ratio_values[0] else "Giảm"
                ws.cell(row=row, column=col).value = trend
                if trend == "Tăng":
                    ws.cell(row=row, column=col).fill = self.fill_trend_up
                else:
                    ws.cell(row=row, column=col).fill = self.fill_trend_down
                ws.cell(row=row, column=col).border = self.border_thin
                col += 1
                
                # Đánh giá
                evaluation = self.evaluate_trend(ratio_name, ratio_values)
                ws.cell(row=row, column=col).value = evaluation
                ws.cell(row=row, column=col).border = self.border_thin
                
            row += 1
            
        # Định dạng cột
        ws.column_dimensions['A'].width = 20
        for i in range(2, 8):
            ws.column_dimensions[chr(64 + i)].width = 15
            
        print("✅ Hoàn thành sheet Phân tích Xu hướng")
        
    def create_growth_analysis_sheet(self):
        """Tạo sheet phân tích tăng trưởng"""
        print("🌱 Đang tạo sheet Phân tích Tăng trưởng...")
        
        ws = self.wb.create_sheet("Phân Tích Tăng Trưởng")
        self.sheets['growth'] = ws
        
        # Header
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "PHÂN TÍCH TỐC ĐỘ TĂNG TRƯỞNG"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Bảng tốc độ tăng trưởng hàng năm
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "TỐC ĐỘ TĂNG TRƯỞNG HÀNG NĂM (%)"
        cell.font = self.font_title
        cell.fill = self.fill_current
        row += 1
        
        # Header
        headers = ['Chỉ tiêu', '2022-2023', '2023-2024', 'Trung bình', 'Xu hướng', 'Dự báo 2025']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_current
            cell.border = self.border_thin
        row += 1
        
        # Các khoản mục cần tính tăng trưởng
        growth_items = [
            ('Tổng tài sản', 'total_assets'),
            ('Tài sản ngắn hạn', 'current_assets'), 
            ('Vốn chủ sở hữu', 'total_equity'),
            ('Doanh thu', 'revenue'),
            ('Lợi nhuận sau thuế', 'net_income'),
            ('Hàng tồn kho', 'inventory'),
            ('Phải thu khách hàng', 'accounts_receivable')
        ]
        
        for item_name, data_key in growth_items:
            ws[f'A{row}'] = item_name
            ws[f'A{row}'].font = self.font_bold
            
            # Tính tăng trưởng từng năm
            periods = self.periods_list
            growth_rates = []
            
            for i in range(1, len(periods)):
                prev_value = self.period_data[periods[i-1]][data_key]
                curr_value = self.period_data[periods[i]][data_key]
                growth_rate = ((curr_value - prev_value) / prev_value) * 100
                growth_rates.append(growth_rate)
                
                ws.cell(row=row, column=i+1).value = growth_rate
                ws.cell(row=row, column=i+1).number_format = '0.00%'
                ws.cell(row=row, column=i+1).border = self.border_thin
                
                if growth_rate > 0:
                    ws.cell(row=row, column=i+1).fill = self.fill_trend_up
                else:
                    ws.cell(row=row, column=i+1).fill = self.fill_trend_down
                    
            # Tăng trưởng trung bình
            avg_growth = sum(growth_rates) / len(growth_rates)
            ws.cell(row=row, column=len(periods)+1).value = avg_growth
            ws.cell(row=row, column=len(periods)+1).number_format = '0.00%'
            ws.cell(row=row, column=len(periods)+1).border = self.border_thin
            
            # Xu hướng
            trend = "Tăng tốc" if growth_rates[-1] > growth_rates[0] else "Chậm lại"
            ws.cell(row=row, column=len(periods)+2).value = trend
            ws.cell(row=row, column=len(periods)+2).border = self.border_thin
            
            # Dự báo năm tiếp theo (công thức Excel)
            last_value = self.period_data[periods[-1]][data_key]
            forecast_formula = f"={last_value}*(1+{avg_growth/100})"
            ws.cell(row=row, column=len(periods)+3).value = forecast_formula
            ws.cell(row=row, column=len(periods)+3).number_format = '#,##0'
            ws.cell(row=row, column=len(periods)+3).border = self.border_thin
            
            row += 1
            
        # Định dạng cột
        ws.column_dimensions['A'].width = 25
        for i in range(2, 7):
            ws.column_dimensions[chr(64 + i)].width = 15
            
        print("✅ Hoàn thành sheet Phân tích Tăng trưởng")
        
    def create_ratio_comparison_sheet(self):
        """Tạo sheet so sánh chỉ số"""
        print("📊 Đang tạo sheet So sánh Chỉ số...")
        
        ws = self.wb.create_sheet("So Sánh Chỉ Số")
        self.sheets['ratios'] = ws
        
        # Header
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = "SO SÁNH CHỈ SỐ TÀI CHÍNH QUA CÁC KỲ"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Bảng so sánh chi tiết
        row = 3
        
        # 1. Nhóm chỉ số thanh khoản
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "1. CHỈ SỐ THANH KHOẢN"
        cell.font = self.font_title
        cell.fill = self.fill_current
        row += 1
        
        self.create_ratio_table(ws, row, [
            ('Current Ratio', 'current_assets', 'current_liabilities'),
            ('Quick Ratio', 'current_assets-inventory', 'current_liabilities'),
        ])
        
        print("✅ Hoàn thành sheet So sánh Chỉ số")
        
    def create_ratio_table(self, ws, start_row, ratios):
        """Tạo bảng chỉ số với công thức động"""
        headers = ['Chỉ số'] + [f'Năm {p}' for p in self.periods_list] + ['Thay đổi', 'Đánh giá']
        
        row = start_row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_current
            cell.border = self.border_thin
        row += 1
        
        for ratio_name, numerator, denominator in ratios:
            ws[f'A{row}'] = ratio_name
            ws[f'A{row}'].font = self.font_bold
            
            col = 2
            for period in self.periods_list:
                # Tạo công thức Excel động
                if '-' in numerator:
                    parts = numerator.split('-')
                    num_formula = f"={parts[0]}_{period}-{parts[1]}_{period}"
                else:
                    num_formula = f"={numerator}_{period}"
                    
                den_formula = f"={denominator}_{period}"
                ratio_formula = f"={num_formula}/{den_formula}"
                
                ws.cell(row=row, column=col).value = ratio_formula
                ws.cell(row=row, column=col).number_format = '0.00'
                ws.cell(row=row, column=col).border = self.border_thin
                col += 1
                
            row += 1
            
    def create_forecast_sheet(self):
        """Tạo sheet dự báo"""
        print("🔮 Đang tạo sheet Dự báo...")
        
        ws = self.wb.create_sheet("Dự Báo Tài Chính")
        self.sheets['forecast'] = ws
        
        # Header
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "DỰ BÁO TÀI CHÍNH DỰA TRÊN XU HƯỚNG"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Bảng dự báo
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "DỰ BÁO CÁC CHỈ TIÊU CHÍNH NĂM 2025"
        cell.font = self.font_title
        cell.fill = self.fill_current
        row += 1
        
        # Header
        headers = ['Chỉ tiêu', 'Năm 2024', 'Tăng trưởng TB', 'Dự báo 2025', 'Kịch bản lạc quan', 'Kịch bản thận trọng']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_current
            cell.border = self.border_thin
        row += 1
        
        # Dự báo các khoản mục chính
        forecast_items = [
            ('Doanh thu', 'revenue'),
            ('Lợi nhuận sau thuế', 'net_income'),
            ('Tổng tài sản', 'total_assets'),
            ('Vốn chủ sở hữu', 'total_equity')
        ]
        
        for item_name, data_key in forecast_items:
            ws[f'A{row}'] = item_name
            ws[f'A{row}'].font = self.font_bold
            
            # Giá trị năm 2024
            value_2024 = self.period_data['2024'][data_key]
            ws.cell(row=row, column=2).value = value_2024
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # Tính tăng trưởng trung bình
            values = [self.period_data[p][data_key] for p in self.periods_list]
            avg_growth = ((values[-1] / values[0]) ** (1/(len(values)-1)) - 1) * 100
            ws.cell(row=row, column=3).value = avg_growth
            ws.cell(row=row, column=3).number_format = '0.00%'
            
            # Dự báo cơ sở (công thức Excel)
            forecast_formula = f"=B{row}*(1+C{row})"
            ws.cell(row=row, column=4).value = forecast_formula
            ws.cell(row=row, column=4).number_format = '#,##0'
            
            # Kịch bản lạc quan (+20% growth)
            optimistic_formula = f"=B{row}*(1+C{row}*1.2)"
            ws.cell(row=row, column=5).value = optimistic_formula
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).fill = self.fill_trend_up
            
            # Kịch bản thận trọng (-20% growth)
            conservative_formula = f"=B{row}*(1+C{row}*0.8)"
            ws.cell(row=row, column=6).value = conservative_formula
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).fill = self.fill_previous
            
            row += 1
            
        # Định dạng cột
        ws.column_dimensions['A'].width = 25
        for i in range(2, 7):
            ws.column_dimensions[chr(64 + i)].width = 18
            
        print("✅ Hoàn thành sheet Dự báo")
        
    def create_named_range(self, name, cell_range):
        """Tạo named range trong workbook"""
        try:
            defined_name = DefinedName(name, attr_text=cell_range)
            if hasattr(self.wb, 'defined_names'):
                self.wb.defined_names[name] = defined_name
        except Exception as e:
            print(f"⚠️  Lỗi tạo named range {name}: {e}")
            
    def evaluate_trend(self, ratio_name, values):
        """Đánh giá xu hướng chỉ số"""
        if len(values) < 2:
            return "Không đủ dữ liệu"
            
        trend = values[-1] - values[0]
        
        if ratio_name in ['Current Ratio', 'ROA (%)', 'ROE (%)', 'Asset Turnover']:
            # Các chỉ số càng cao càng tốt
            if trend > 0:
                return "Cải thiện tích cực"
            elif trend < 0:
                return "Cần chú ý"
            else:
                return "Ổn định"
        else:
            return "Biến động"
            
    def save_workbook(self, filename=None):
        """Lưu workbook"""
        if not filename:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"phan_tich_nhieu_ky_{timestamp}.xlsx"
            
        try:
            self.wb.save(filename)
            print(f"✅ Đã lưu file: {filename}")
            return filename
        except Exception as e:
            print(f"❌ Lỗi lưu file: {e}")
            return None
            
    def generate_complete_analysis(self):
        """Tạo hoàn chỉnh hệ thống phân tích nhiều kỳ"""
        print("🚀 Bắt đầu tạo hệ thống phân tích nhiều kỳ...")
        
        # Tạo workbook
        self.create_multi_period_workbook()
        
        # Lưu file
        filename = self.save_workbook()
        
        if filename:
            print(f"\n🎉 HOÀN THÀNH HỆ THỐNG PHÂN TÍCH NHIỀU KỲ!")
            print(f"📁 File: {filename}")
            print(f"📊 Sheets: {len(self.sheets)}")
            print(f"📈 Periods: {len(self.periods_list)}")
            
            print("\n📋 Các sheet đã tạo:")
            for name, sheet in self.sheets.items():
                print(f"   ✓ {sheet.title}")
                
            print("\n🔍 Tính năng chính:")
            print(f"   ✓ So sánh {len(self.periods_list)} kỳ báo cáo")
            print(f"   ✓ Phân tích xu hướng tăng trưởng")
            print(f"   ✓ Dự báo tài chính tự động")
            print(f"   ✓ Công thức Excel động cho tất cả tính toán")
                    
        return filename

# Test và chạy
if __name__ == "__main__":
    analyzer = MultiPeriodAnalyzer(periods=3)
    filename = analyzer.generate_complete_analysis()
    
    if filename:
        print(f"\n💡 Hướng dẫn sử dụng:")
        print(f"1. Thay đổi dữ liệu trong sheet 'Bảng Cân Đối Nhiều Kỳ'")
        print(f"2. Tất cả báo cáo khác sẽ tự động cập nhật")
        print(f"3. Xem xu hướng phát triển qua các kỳ")
        print(f"4. Sử dụng dự báo để lập kế hoạch tài chính")