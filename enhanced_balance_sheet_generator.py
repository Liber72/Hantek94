"""
Enhanced Balance Sheet Generator với Named Ranges cho Hệ thống Báo cáo Tài chính Động
============================================================================

Tạo bảng cân đối kế toán với hệ thống named ranges hoàn chỉnh theo chuẩn kế toán Việt Nam
Hỗ trợ công thức Excel động cho các báo cáo tài chính

Tác giả: Hệ thống Phân tích Tài chính Động
Chuẩn: VAS/Circular 200/2014/TT-BTC
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
import datetime
import json
import os

class EnhancedBalanceSheetGenerator:
    def __init__(self):
        self.wb = None
        self.ws_balance = None
        self.ws_mapping = None
        self.ws_income = None
        self.named_ranges = {}
        self.accounting_codes = {}
        
        # Thiết lập style
        self.setup_styles()
        
        # Dữ liệu mẫu theo chuẩn kế toán Việt Nam
        self.setup_sample_data()
        
    def setup_styles(self):
        """Thiết lập các style Excel chuyên nghiệp"""
        self.font_header = Font(name='Times New Roman', size=14, bold=True)
        self.font_title = Font(name='Times New Roman', size=12, bold=True)
        self.font_normal = Font(name='Times New Roman', size=11)
        self.font_bold = Font(name='Times New Roman', size=11, bold=True)
        
        # Màu sắc
        self.fill_header = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.fill_section = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        self.fill_subsection = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Border
        self.border_thick = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thick'), bottom=Side(style='thick')
        )
        self.border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
    def setup_sample_data(self):
        """Thiết lập dữ liệu mẫu theo chuẩn kế toán Việt Nam"""
        # Dữ liệu được thiết kế để cân đối chính xác
        self.balance_sheet_data = {
            # TÀI SẢN
            'assets': {
                'current_assets': {
                    'cash_and_equivalents': {'code': '111', 'value': 15000000000, 'name': 'Tiền và tương đương tiền'},
                    'short_term_investments': {'code': '121', 'value': 8000000000, 'name': 'Đầu tư tài chính ngắn hạn'},
                    'accounts_receivable': {'code': '131', 'value': 12000000000, 'name': 'Phải thu khách hàng'},
                    'inventory': {'code': '141', 'value': 18000000000, 'name': 'Hàng tồn kho'},
                    'prepaid_expenses': {'code': '151', 'value': 2000000000, 'name': 'Chi phí trả trước ngắn hạn'},
                    'other_current_assets': {'code': '161', 'value': 3000000000, 'name': 'Tài sản ngắn hạn khác'}
                    # Tổng: 58,000,000,000
                },
                'non_current_assets': {
                    'long_term_receivables': {'code': '211', 'value': 5000000000, 'name': 'Phải thu dài hạn'},
                    'fixed_assets': {'code': '221', 'value': 45000000000, 'name': 'Tài sản cố định hữu hình'},
                    'intangible_assets': {'code': '231', 'value': 8000000000, 'name': 'Tài sản cố định vô hình'},
                    'long_term_investments': {'code': '241', 'value': 12000000000, 'name': 'Đầu tư tài chính dài hạn'},
                    'other_non_current_assets': {'code': '251', 'value': 2000000000, 'name': 'Tài sản dài hạn khác'}
                    # Tổng: 72,000,000,000
                }
                # Tổng tài sản: 130,000,000,000
            },
            # NỢ PHẢI TRẢ
            'liabilities': {
                'current_liabilities': {
                    'accounts_payable': {'code': '311', 'value': 10000000000, 'name': 'Phải trả người bán'},
                    'short_term_loans': {'code': '321', 'value': 15000000000, 'name': 'Vay và nợ ngắn hạn'},
                    'accrued_expenses': {'code': '331', 'value': 5000000000, 'name': 'Chi phí phải trả'},
                    'taxes_payable': {'code': '341', 'value': 3000000000, 'name': 'Thuế và phí phải nộp'},
                    'other_current_liabilities': {'code': '351', 'value': 2000000000, 'name': 'Nợ ngắn hạn khác'}
                    # Tổng: 35,000,000,000
                },
                'non_current_liabilities': {
                    'long_term_loans': {'code': '411', 'value': 25000000000, 'name': 'Vay và nợ dài hạn'},
                    'provisions': {'code': '421', 'value': 3000000000, 'name': 'Dự phòng dài hạn'},
                    'other_non_current_liabilities': {'code': '431', 'value': 2000000000, 'name': 'Nợ dài hạn khác'}
                    # Tổng: 30,000,000,000
                }
                # Tổng nợ: 65,000,000,000
            },
            # VỐN CHỦ SỞ HỮU
            'equity': {
                'share_capital': {'code': '511', 'value': 50000000000, 'name': 'Vốn điều lệ'},
                'capital_surplus': {'code': '521', 'value': 5000000000, 'name': 'Thặng dư vốn cổ phần'},
                'retained_earnings': {'code': '531', 'value': 10000000000, 'name': 'Lợi nhuận sau thuế chưa phân phối'},
                'other_equity': {'code': '541', 'value': 0, 'name': 'Nguồn vốn chủ sở hữu khác'}
                # Tổng vốn: 65,000,000,000
            }
            # Tổng nguồn vốn: 65,000,000,000 + 65,000,000,000 = 130,000,000,000 ✓
        }
        
        # Dữ liệu báo cáo kết quả kinh doanh
        self.income_statement_data = {
            'revenue': {'code': '511', 'value': 80000000000, 'name': 'Doanh thu thuần'},
            'cost_of_goods_sold': {'code': '621', 'value': 50000000000, 'name': 'Giá vốn hàng bán'},
            'gross_profit': {'code': '631', 'value': 30000000000, 'name': 'Lợi nhuận gộp'},
            'operating_expenses': {'code': '641', 'value': 20000000000, 'name': 'Chi phí bán hàng và quản lý'},
            'operating_income': {'code': '651', 'value': 10000000000, 'name': 'Lợi nhuận từ hoạt động kinh doanh'},
            'financial_income': {'code': '661', 'value': 2000000000, 'name': 'Thu nhập tài chính'},
            'financial_expenses': {'code': '671', 'value': 1500000000, 'name': 'Chi phí tài chính'},
            'other_income': {'code': '681', 'value': 500000000, 'name': 'Thu nhập khác'},
            'other_expenses': {'code': '691', 'value': 200000000, 'name': 'Chi phí khác'},
            'pre_tax_income': {'code': '701', 'value': 10800000000, 'name': 'Lợi nhuận trước thuế'},
            'tax_expense': {'code': '711', 'value': 2160000000, 'name': 'Chi phí thuế TNDN'},
            'net_income': {'code': '721', 'value': 8640000000, 'name': 'Lợi nhuận sau thuế'}
        }
        
    def create_balance_sheet(self):
        """Tạo sheet bảng cân đối kế toán với named ranges"""
        print("🏗️  Đang tạo bảng cân đối kế toán...")
        
        self.wb = openpyxl.Workbook()
        self.ws_balance = self.wb.active
        self.ws_balance.title = "Bảng Cân Đối Kế Toán"
        
        # Header
        self.create_balance_sheet_header()
        
        # Tài sản
        row = 6
        row = self.create_assets_section(row)
        
        # Nợ phải trả
        row += 2
        row = self.create_liabilities_section(row)
        
        # Vốn chủ sở hữu
        row += 2
        row = self.create_equity_section(row)
        
        # Kiểm tra cân đối
        self.create_balance_check(row + 2)
        
        # Định dạng cột
        self.format_balance_sheet_columns()
        
        print("✅ Hoàn thành bảng cân đối kế toán với named ranges")
        
    def create_balance_sheet_header(self):
        """Tạo header cho bảng cân đối kế toán"""
        # Tiêu đề chính
        self.ws_balance.merge_cells('A1:D1')
        cell = self.ws_balance['A1']
        cell.value = "BẢNG CÂN ĐỐI KẾ TOÁN"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Thông tin công ty
        self.ws_balance.merge_cells('A2:D2')
        cell = self.ws_balance['A2']
        cell.value = f"Tại ngày: {datetime.date.today().strftime('%d/%m/%Y')}"
        cell.font = self.font_title
        cell.alignment = Alignment(horizontal='center')
        
        # Header cột
        headers = ['Chỉ tiêu', 'Mã số', 'Thuyết minh', 'Số cuối kỳ (VND)']
        for col, header in enumerate(headers, 1):
            cell = self.ws_balance.cell(row=4, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_section
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    def create_assets_section(self, start_row):
        """Tạo phần tài sản"""
        row = start_row
        
        # Header Tài sản
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "TÀI SẢN"
        cell.font = self.font_bold
        cell.fill = self.fill_section
        cell.border = self.border_thin
        row += 1
        
        # Tài sản ngắn hạn
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "A. TÀI SẢN NGẮN HẠN"
        cell.font = self.font_bold
        cell.fill = self.fill_subsection
        row += 1
        
        current_assets_start = row
        current_assets_total = 0
        
        for key, item in self.balance_sheet_data['assets']['current_assets'].items():
            self.ws_balance[f'A{row}'] = f"   {item['name']}"
            self.ws_balance[f'B{row}'] = item['code']
            self.ws_balance[f'C{row}'] = ""
            self.ws_balance[f'D{row}'] = item['value']
            
            # Tạo named range cho từng item
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f"'Bảng Cân Đối Kế Toán'!D{row}")
            
            current_assets_total += item['value']
            row += 1
            
        # Tổng tài sản ngắn hạn
        self.ws_balance[f'A{row}'] = "Tổng tài sản ngắn hạn"
        self.ws_balance[f'B{row}'] = "100"
        self.ws_balance[f'D{row}'] = current_assets_total
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        self.create_named_range('CurrentAssets', f"'Bảng Cân Đối Kế Toán'!D{row}")
        row += 2
        
        # Tài sản dài hạn
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "B. TÀI SẢN DÀI HẠN"
        cell.font = self.font_bold
        cell.fill = self.fill_subsection
        row += 1
        
        non_current_assets_total = 0
        
        for key, item in self.balance_sheet_data['assets']['non_current_assets'].items():
            self.ws_balance[f'A{row}'] = f"   {item['name']}"
            self.ws_balance[f'B{row}'] = item['code']
            self.ws_balance[f'C{row}'] = ""
            self.ws_balance[f'D{row}'] = item['value']
            
            # Tạo named range
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f'D{row}')
            
            non_current_assets_total += item['value']
            row += 1
            
        # Tổng tài sản dài hạn
        self.ws_balance[f'A{row}'] = "Tổng tài sản dài hạn"
        self.ws_balance[f'B{row}'] = "200"
        self.ws_balance[f'D{row}'] = non_current_assets_total
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        self.create_named_range('NonCurrentAssets', f'D{row}')
        row += 1
        
        # Tổng tài sản
        total_assets = current_assets_total + non_current_assets_total
        self.ws_balance[f'A{row}'] = "TỔNG CỘNG TÀI SẢN"
        self.ws_balance[f'B{row}'] = "270"
        self.ws_balance[f'D{row}'] = total_assets
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold, self.fill_section)
        self.create_named_range('TotalAssets', f"'Bảng Cân Đối Kế Toán'!D{row}")
        
        return row + 1
        
    def create_liabilities_section(self, start_row):
        """Tạo phần nợ phải trả"""
        row = start_row
        
        # Header Nợ phải trả
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "NGUỒN VỐN"
        cell.font = self.font_bold
        cell.fill = self.fill_section
        cell.border = self.border_thin
        row += 1
        
        # Nợ ngắn hạn
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "C. NỢ PHẢI TRẢ"
        cell.font = self.font_bold
        cell.fill = self.fill_subsection
        row += 1
        
        self.ws_balance[f'A{row}'] = "I. Nợ ngắn hạn"
        self.ws_balance[f'B{row}'] = "300"
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        row += 1
        
        current_liabilities_total = 0
        
        for key, item in self.balance_sheet_data['liabilities']['current_liabilities'].items():
            self.ws_balance[f'A{row}'] = f"   {item['name']}"
            self.ws_balance[f'B{row}'] = item['code']
            self.ws_balance[f'C{row}'] = ""
            self.ws_balance[f'D{row}'] = item['value']
            
            # Tạo named range
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f'D{row}')
            
            current_liabilities_total += item['value']
            row += 1
            
        # Tổng nợ ngắn hạn
        self.ws_balance[f'A{row}'] = "Tổng nợ ngắn hạn"
        self.ws_balance[f'B{row}'] = "300"
        self.ws_balance[f'D{row}'] = current_liabilities_total
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        self.create_named_range('CurrentLiabilities', f'D{row}')
        row += 2
        
        # Nợ dài hạn
        self.ws_balance[f'A{row}'] = "II. Nợ dài hạn"
        self.ws_balance[f'B{row}'] = "400"
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        row += 1
        
        non_current_liabilities_total = 0
        
        for key, item in self.balance_sheet_data['liabilities']['non_current_liabilities'].items():
            self.ws_balance[f'A{row}'] = f"   {item['name']}"
            self.ws_balance[f'B{row}'] = item['code']
            self.ws_balance[f'C{row}'] = ""
            self.ws_balance[f'D{row}'] = item['value']
            
            # Tạo named range
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f'D{row}')
            
            non_current_liabilities_total += item['value']
            row += 1
            
        # Tổng nợ dài hạn
        self.ws_balance[f'A{row}'] = "Tổng nợ dài hạn"
        self.ws_balance[f'B{row}'] = "400"
        self.ws_balance[f'D{row}'] = non_current_liabilities_total
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold)
        self.create_named_range('NonCurrentLiabilities', f'D{row}')
        row += 1
        
        # Tổng nợ phải trả
        total_liabilities = current_liabilities_total + non_current_liabilities_total
        self.ws_balance[f'A{row}'] = "TỔNG CỘNG NỢ PHẢI TRẢ"
        self.ws_balance[f'B{row}'] = "430"
        self.ws_balance[f'D{row}'] = total_liabilities
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold, self.fill_section)
        self.create_named_range('TotalLiabilities', f'D{row}')
        
        return row
        
    def create_equity_section(self, start_row):
        """Tạo phần vốn chủ sở hữu"""
        row = start_row
        
        # Header vốn chủ sở hữu
        self.ws_balance.merge_cells(f'A{row}:D{row}')
        cell = self.ws_balance[f'A{row}']
        cell.value = "D. VỐN CHỦ SỞ HỮU"
        cell.font = self.font_bold
        cell.fill = self.fill_subsection
        row += 1
        
        total_equity = 0
        
        for key, item in self.balance_sheet_data['equity'].items():
            self.ws_balance[f'A{row}'] = f"   {item['name']}"
            self.ws_balance[f'B{row}'] = item['code']
            self.ws_balance[f'C{row}'] = ""
            self.ws_balance[f'D{row}'] = item['value']
            
            # Tạo named range
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f'D{row}')
            
            total_equity += item['value']
            row += 1
            
        # Tổng vốn chủ sở hữu
        self.ws_balance[f'A{row}'] = "TỔNG CỘNG VỐN CHỦ SỞ HỮU"
        self.ws_balance[f'B{row}'] = "440"
        self.ws_balance[f'D{row}'] = total_equity
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold, self.fill_section)
        self.create_named_range('TotalEquity', f'D{row}')
        row += 1
        
        # Tổng nguồn vốn
        total_liabilities = sum(item['value'] for section in self.balance_sheet_data['liabilities'].values() 
                               for item in section.values())
        
        self.ws_balance[f'A{row}'] = "TỔNG CỘNG NGUỒN VỐN"
        self.ws_balance[f'B{row}'] = "440"
        self.ws_balance[f'D{row}'] = total_liabilities + total_equity
        self.apply_cell_style(self.ws_balance[f'A{row}:D{row}'], self.font_bold, self.fill_header)
        self.create_named_range('TotalLiabilitiesAndEquity', f'D{row}')
        
        return row
        
    def create_balance_check(self, row):
        """Tạo phần kiểm tra cân đối"""
        self.ws_balance[f'A{row}'] = "KIỂM TRA CÂN ĐỐI:"
        self.ws_balance[f'A{row+1}'] = "Tổng Tài sản ="
        self.ws_balance[f'B{row+1}'] = "=TotalAssets"
        self.ws_balance[f'A{row+2}'] = "Tổng Nguồn vốn ="
        self.ws_balance[f'B{row+2}'] = "=TotalLiabilitiesAndEquity"
        self.ws_balance[f'A{row+3}'] = "Chênh lệch ="
        self.ws_balance[f'B{row+3}'] = "=TotalAssets-TotalLiabilitiesAndEquity"
        
        # Định dạng
        for r in range(row, row+4):
            self.apply_cell_style(self.ws_balance[f'A{r}:B{r}'], self.font_bold)
            
    def create_income_statement_sheet(self):
        """Tạo sheet báo cáo kết quả kinh doanh"""
        print("📊 Đang tạo báo cáo kết quả kinh doanh...")
        
        self.ws_income = self.wb.create_sheet("Báo Cáo Kết Quả Kinh Doanh")
        
        # Header
        self.ws_income.merge_cells('A1:D1')
        cell = self.ws_income['A1']
        cell.value = "BÁO CÁO KẾT QUẢ KINH DOANH"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Thông tin
        self.ws_income.merge_cells('A2:D2')
        cell = self.ws_income['A2']
        cell.value = f"Năm kết thúc ngày: {datetime.date.today().strftime('%d/%m/%Y')}"
        cell.font = self.font_title
        cell.alignment = Alignment(horizontal='center')
        
        # Header cột
        headers = ['Chỉ tiêu', 'Mã số', 'Thuyết minh', 'Năm nay (VND)']
        for col, header in enumerate(headers, 1):
            cell = self.ws_income.cell(row=4, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_section
            cell.border = self.border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Dữ liệu
        row = 5
        for key, item in self.income_statement_data.items():
            self.ws_income[f'A{row}'] = item['name']
            self.ws_income[f'B{row}'] = item['code']
            self.ws_income[f'C{row}'] = ""
            self.ws_income[f'D{row}'] = item['value']
            
            # Tạo named range
            range_name = self.camel_case(key)
            self.create_named_range(range_name, f"'Báo Cáo Kết Quả Kinh Doanh'!D{row}")
            
            row += 1
            
        # Định dạng cột
        self.format_income_statement_columns()
        
        print("✅ Hoàn thành báo cáo kết quả kinh doanh với named ranges")
        
    def create_mapping_sheet(self):
        """Tạo sheet mapping với các công thức ví dụ"""
        print("🗺️  Đang tạo sheet mapping và ví dụ công thức...")
        
        self.ws_mapping = self.wb.create_sheet("Mapping và Công thức")
        
        # Header
        self.ws_mapping.merge_cells('A1:E1')
        cell = self.ws_mapping['A1']
        cell.value = "HỆ THỐNG MAPPING VÀ CÔNG THỨC ĐỘNG"
        cell.font = self.font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = self.fill_header
        
        # Bảng named ranges
        row = 3
        self.ws_mapping[f'A{row}'] = "DANH SÁCH NAMED RANGES"
        self.apply_cell_style(self.ws_mapping[f'A{row}:E{row}'], self.font_bold, self.fill_section)
        row += 1
        
        headers = ['Named Range', 'Mô tả', 'Cell/Range', 'Giá trị', 'Công thức ví dụ']
        for col, header in enumerate(headers, 1):
            cell = self.ws_mapping.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_subsection
            
        row += 1
        
        # Danh sách named ranges với ví dụ công thức
        mapping_examples = [
            ('TotalAssets', 'Tổng tài sản', 'Sheet1!D26', '=TotalAssets', 'ROA = NetIncome/TotalAssets'),
            ('CurrentAssets', 'Tài sản ngắn hạn', 'Sheet1!D12', '=CurrentAssets', 'Current Ratio = CurrentAssets/CurrentLiabilities'),
            ('CurrentLiabilities', 'Nợ ngắn hạn', 'Sheet1!D35', '=CurrentLiabilities', 'Quick Ratio = (CurrentAssets-Inventory)/CurrentLiabilities'),
            ('TotalEquity', 'Vốn chủ sở hữu', 'Sheet1!D45', '=TotalEquity', 'ROE = NetIncome/TotalEquity'),
            ('Revenue', 'Doanh thu', 'Sheet2!D5', '=Revenue', 'Asset Turnover = Revenue/TotalAssets'),
            ('NetIncome', 'Lợi nhuận sau thuế', 'Sheet2!D17', '=NetIncome', 'Profit Margin = NetIncome/Revenue'),
            ('Inventory', 'Hàng tồn kho', 'Sheet1!D9', '=Inventory', 'Inventory Turnover = COGS/Inventory'),
            ('TotalLiabilities', 'Tổng nợ', 'Sheet1!D41', '=TotalLiabilities', 'Debt to Assets = TotalLiabilities/TotalAssets')
        ]
        
        for range_name, description, cell_ref, formula, example in mapping_examples:
            self.ws_mapping[f'A{row}'] = range_name
            self.ws_mapping[f'B{row}'] = description
            self.ws_mapping[f'C{row}'] = cell_ref
            self.ws_mapping[f'D{row}'] = formula
            self.ws_mapping[f'E{row}'] = example
            row += 1
            
        # Bảng công thức chỉ số tài chính
        row += 2
        self.ws_mapping[f'A{row}'] = "CÁC CÔNG THỨC CHỈ SỐ TÀI CHÍNH"
        self.apply_cell_style(self.ws_mapping[f'A{row}:E{row}'], self.font_bold, self.fill_section)
        row += 1
        
        headers = ['Chỉ số', 'Công thức Excel', 'Ý nghĩa', 'Chuẩn đánh giá', 'Công thức Python tương đương']
        for col, header in enumerate(headers, 1):
            cell = self.ws_mapping.cell(row=row, column=col)
            cell.value = header
            cell.font = self.font_bold
            cell.fill = self.fill_subsection
            
        row += 1
        
        financial_ratios = [
            ('Current Ratio', '=CurrentAssets/CurrentLiabilities', 'Khả năng thanh toán ngắn hạn', '>2.0: Tốt', 'current_assets / current_liabilities'),
            ('Quick Ratio', '=(CurrentAssets-Inventory)/CurrentLiabilities', 'Khả năng thanh toán tức thời', '>1.0: Tốt', '(current_assets - inventory) / current_liabilities'),
            ('ROA (%)', '=NetIncome/TotalAssets*100', 'Hiệu quả sử dụng tài sản', '>5%: Tốt', 'net_income / total_assets * 100'),
            ('ROE (%)', '=NetIncome/TotalEquity*100', 'Hiệu quả sử dụng vốn', '>15%: Tốt', 'net_income / total_equity * 100'),
            ('Debt to Assets', '=TotalLiabilities/TotalAssets', 'Tỷ lệ nợ trên tài sản', '<0.6: Tốt', 'total_liabilities / total_assets'),
            ('Asset Turnover', '=Revenue/TotalAssets', 'Hiệu quả quay vòng tài sản', '>1.0: Tốt', 'revenue / total_assets'),
            ('Inventory Turnover', '=CostOfGoodsSold/Inventory', 'Hiệu quả quay vòng hàng tồn', '>6: Tốt', 'cogs / inventory'),
            ('Profit Margin (%)', '=NetIncome/Revenue*100', 'Tỷ suất lợi nhuận', '>10%: Tốt', 'net_income / revenue * 100')
        ]
        
        for ratio_name, formula, meaning, standard, python_equiv in financial_ratios:
            self.ws_mapping[f'A{row}'] = ratio_name
            self.ws_mapping[f'B{row}'] = formula
            self.ws_mapping[f'C{row}'] = meaning
            self.ws_mapping[f'D{row}'] = standard
            self.ws_mapping[f'E{row}'] = python_equiv
            row += 1
            
        # Định dạng
        self.format_mapping_columns()
        
        print("✅ Hoàn thành sheet mapping với công thức ví dụ")
        
    def create_named_range(self, name, cell_range):
        """Tạo named range trong workbook"""
        try:
            # Tạo DefinedName object đúng cách cho openpyxl
            defined_name = DefinedName(name, attr_text=cell_range)
            
            # Thêm vào workbook defined names
            if hasattr(self.wb, 'defined_names'):
                # Xóa range cũ nếu tồn tại
                if name in self.wb.defined_names:
                    del self.wb.defined_names[name]
                    
                # Thêm range mới
                self.wb.defined_names[name] = defined_name
                
                # Lưu vào dict để tracking
                self.named_ranges[name] = cell_range
                print(f"✅ Tạo named range: {name} -> {cell_range}")
            else:
                # Fallback cho các phiên bản openpyxl cũ
                self.named_ranges[name] = cell_range
                print(f"⚠️  Lưu named range (không thêm vào workbook): {name} -> {cell_range}")
            
        except Exception as e:
            print(f"⚠️  Lỗi tạo named range {name}: {e}")
            # Lưu thông tin để sử dụng sau
            self.named_ranges[name] = cell_range
            
    def camel_case(self, text):
        """Chuyển text thành camelCase cho named range"""
        words = text.split('_')
        return words[0].lower() + ''.join(word.capitalize() for word in words[1:])
        
    def apply_cell_style(self, cell_range, font=None, fill=None, border=None):
        """Áp dụng style cho range cells"""
        if isinstance(cell_range, str):
            # Nếu là string thì convert thành range
            start_cell, end_cell = cell_range.split(':')
            cell_range = self.ws_balance[cell_range]
            
        if hasattr(cell_range, '__iter__'):
            # Nếu là range
            for row in cell_range:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if border:
                        cell.border = border or self.border_thin
        else:
            # Nếu là single cell
            if font:
                cell_range.font = font
            if fill:
                cell_range.fill = fill
            if border:
                cell_range.border = border or self.border_thin
                
    def format_balance_sheet_columns(self):
        """Định dạng cột cho bảng cân đối"""
        # Độ rộng cột
        self.ws_balance.column_dimensions['A'].width = 40
        self.ws_balance.column_dimensions['B'].width = 10
        self.ws_balance.column_dimensions['C'].width = 15
        self.ws_balance.column_dimensions['D'].width = 20
        
        # Định dạng số
        for row in self.ws_balance.iter_rows():
            for cell in row:
                if cell.column == 4 and cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    
    def format_income_statement_columns(self):
        """Định dạng cột cho báo cáo kết quả kinh doanh"""
        # Độ rộng cột
        self.ws_income.column_dimensions['A'].width = 40
        self.ws_income.column_dimensions['B'].width = 10
        self.ws_income.column_dimensions['C'].width = 15
        self.ws_income.column_dimensions['D'].width = 20
        
        # Định dạng số
        for row in self.ws_income.iter_rows():
            for cell in row:
                if cell.column == 4 and cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    
    def format_mapping_columns(self):
        """Định dạng cột cho sheet mapping"""
        # Độ rộng cột
        self.ws_mapping.column_dimensions['A'].width = 20
        self.ws_mapping.column_dimensions['B'].width = 30
        self.ws_mapping.column_dimensions['C'].width = 15
        self.ws_mapping.column_dimensions['D'].width = 20
        self.ws_mapping.column_dimensions['E'].width = 40
        
    def save_workbook(self, filename=None):
        """Lưu workbook"""
        if not filename:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bang_can_doi_ke_toan_dynamic_{timestamp}.xlsx"
            
        try:
            self.wb.save(filename)
            print(f"✅ Đã lưu file: {filename}")
            
            # Lưu thông tin named ranges
            self.save_named_ranges_info(filename.replace('.xlsx', '_named_ranges.json'))
            
            return filename
        except Exception as e:
            print(f"❌ Lỗi lưu file: {e}")
            return None
            
    def save_named_ranges_info(self, filename):
        """Lưu thông tin named ranges ra file JSON"""
        try:
            info = {
                'created': datetime.datetime.now().isoformat(),
                'total_ranges': len(self.named_ranges),
                'ranges': self.named_ranges,
                'balance_sheet_data': self.balance_sheet_data,
                'income_statement_data': self.income_statement_data
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(info, f, ensure_ascii=False, indent=2)
                
            print(f"✅ Đã lưu thông tin named ranges: {filename}")
            
        except Exception as e:
            print(f"⚠️  Lỗi lưu named ranges info: {e}")
            
    def generate_complete_balance_sheet(self):
        """Tạo hoàn chỉnh bảng cân đối kế toán với named ranges"""
        print("🚀 Bắt đầu tạo hệ thống bảng cân đối kế toán động...")
        
        # Tạo các sheet
        self.create_balance_sheet()
        self.create_income_statement_sheet()
        self.create_mapping_sheet()
        
        # Lưu file
        filename = self.save_workbook()
        
        if filename:
            print(f"\n🎉 HOÀN THÀNH!")
            print(f"📁 File: {filename}")
            print(f"📊 Named Ranges: {len(self.named_ranges)}")
            print(f"📈 Sheets: {len(self.wb.sheetnames)}")
            print("\n📋 Danh sách Named Ranges chính:")
            key_ranges = ['TotalAssets', 'CurrentAssets', 'CurrentLiabilities', 'TotalEquity', 'Revenue', 'NetIncome']
            for range_name in key_ranges:
                if range_name in self.named_ranges:
                    print(f"   ✓ {range_name}: {self.named_ranges[range_name]}")
                    
        return filename

# Test và chạy
if __name__ == "__main__":
    generator = EnhancedBalanceSheetGenerator()
    filename = generator.generate_complete_balance_sheet()
    
    if filename:
        print(f"\n🔍 Hướng dẫn sử dụng:")
        print(f"1. Mở file {filename} trong Excel")
        print(f"2. Các named ranges đã được tạo tự động")
        print(f"3. Sử dụng các công thức như: =CurrentAssets/CurrentLiabilities")
        print(f"4. Thay đổi dữ liệu trong bảng cân đối → các công thức tự động cập nhật")
        print(f"5. Xem sheet 'Mapping và Công thức' để hiểu cách sử dụng")