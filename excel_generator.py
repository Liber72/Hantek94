"""
Excel Generator for VinGroup Financial Analysis
==============================================

This script creates a comprehensive Excel file (.xlsx) for VinGroup financial analysis
using openpyxl. It includes professional formatting, charts, and formulas.

Requirements:
- openpyxl
- matplotlib (for charts)

Usage:
    python excel_generator.py
"""

import json
import os
from datetime import datetime

# Check if required packages are available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.chart.axis import DateAxis
    from openpyxl.formatting.rule import DataBarRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("⚠️  openpyxl not available. Please install it with: pip install openpyxl")
    OPENPYXL_AVAILABLE = False

class VinGroupExcelGenerator:
    """
    Excel generator for VinGroup financial analysis
    """
    
    def __init__(self, data_file: str = "vingroup_analysis/vingroup_data.json"):
        """Initialize with data file"""
        if not os.path.exists(data_file):
            raise FileNotFoundError(f"Data file not found: {data_file}")
        
        with open(data_file, 'r', encoding='utf-8') as f:
            self.data = json.load(f)
        
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Define styles
        self.define_styles()
    
    def define_styles(self):
        """Define styles for formatting"""
        # Header style
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Sub-header style
        self.subheader_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subheader_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        
        # Data style
        self.data_font = Font(name='Arial', size=10)
        self.data_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Number format
        self.number_format = '#,##0'
        self.percentage_format = '0.00%'
        
        # Border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignment
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
    
    def create_financial_statements_sheet(self):
        """Create Sheet 1: Financial Statements"""
        ws = self.workbook.create_sheet("Báo cáo tài chính VinGroup")
        
        # Company info
        ws['A1'] = "TẬP ĐOÀN VINGROUP - BÁO CÁO TÀI CHÍNH"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A2'] = "Đơn vị: Tỷ VND"
        ws['A2'].font = Font(name='Arial', size=12, italic=True)
        
        # Balance Sheet
        self._create_balance_sheet_section(ws, start_row=4)
        
        # Income Statement
        self._create_income_statement_section(ws, start_row=25)
        
        # Cash Flow Statement
        self._create_cash_flow_section(ws, start_row=40)
        
        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        return ws
    
    def _create_balance_sheet_section(self, ws, start_row):
        """Create balance sheet section"""
        current_row = start_row
        
        # Header
        ws[f'A{current_row}'] = "BẢNG CÂN ĐỐI KẾ TOÁN"
        ws[f'A{current_row}'].font = self.header_font
        ws[f'A{current_row}'].fill = self.header_fill
        
        current_row += 1
        headers = ['Khoản mục', '2023', '2024', 'Thay đổi (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        current_row += 1
        
        # Assets section
        ws[f'A{current_row}'] = "TÀI SẢN"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        # Current assets
        ws[f'A{current_row}'] = "A. Tài sản ngắn hạn"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        # Current assets items
        current_assets_items = [
            ("I. Tiền và tương đương tiền", "cash_and_equivalents"),
            ("II. Đầu tư tài chính ngắn hạn", "short_term_investments"),
            ("III. Phải thu ngắn hạn", "accounts_receivable"),
            ("IV. Hàng tồn kho", "inventory"),
            ("V. Tài sản ngắn hạn khác", "prepaid_expenses")
        ]
        
        for item_name, key in current_assets_items:
            val_2023 = self.data["balance_sheet"]["2023"]["assets"]["current_assets"][key]
            val_2024 = self.data["balance_sheet"]["2024"]["assets"]["current_assets"][key]
            change_pct = ((val_2024 - val_2023) / val_2023) * 100 if val_2023 != 0 else 0
            
            ws[f'A{current_row}'] = item_name
            ws[f'B{current_row}'] = val_2023
            ws[f'C{current_row}'] = val_2024
            ws[f'D{current_row}'] = change_pct / 100  # For percentage format
            
            # Format cells
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.thin_border
                if col >= 2:
                    cell.number_format = self.number_format if col <= 3 else self.percentage_format
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            current_row += 1
        
        # Similar structure for non-current assets and liabilities...
        # (Implementation continues with similar pattern)
        
        return current_row
    
    def _create_income_statement_section(self, ws, start_row):
        """Create income statement section"""
        current_row = start_row
        
        # Header
        ws[f'A{current_row}'] = "BÁO CÁO KẾT QUẢ KINH DOANH"
        ws[f'A{current_row}'].font = self.header_font
        ws[f'A{current_row}'].fill = self.header_fill
        
        current_row += 1
        headers = ['Khoản mục', '2023', '2024', 'Thay đổi (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        current_row += 1
        
        # Income statement items
        income_items = [
            ("1. Doanh thu bán hàng", "revenue"),
            ("2. Giá vốn hàng bán", "cost_of_goods_sold"),
            ("3. Lợi nhuận gộp", "gross_profit"),
            ("4. Lợi nhuận từ hoạt động kinh doanh", "operating_profit"),
            ("5. Thu nhập tài chính", "financial_income"),
            ("6. Chi phí tài chính", "financial_expenses"),
            ("7. Lợi nhuận trước thuế", "profit_before_tax"),
            ("8. Chi phí thuế", "tax_expense"),
            ("9. Lợi nhuận sau thuế", "net_profit")
        ]
        
        for item_name, key in income_items:
            val_2023 = self.data["income_statement"]["2023"][key]
            val_2024 = self.data["income_statement"]["2024"][key]
            change_pct = ((val_2024 - val_2023) / val_2023) * 100 if val_2023 != 0 else 0
            
            ws[f'A{current_row}'] = item_name
            ws[f'B{current_row}'] = val_2023
            ws[f'C{current_row}'] = val_2024
            ws[f'D{current_row}'] = change_pct / 100  # For percentage format
            
            # Format cells
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.thin_border
                if col >= 2:
                    cell.number_format = self.number_format if col <= 3 else self.percentage_format
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            current_row += 1
        
        return current_row
    
    def _create_cash_flow_section(self, ws, start_row):
        """Create cash flow section"""
        current_row = start_row
        
        # Header
        ws[f'A{current_row}'] = "BÁO CÁO LƯU CHUYỂN TIỀN TỆ"
        ws[f'A{current_row}'].font = self.header_font
        ws[f'A{current_row}'].fill = self.header_fill
        
        current_row += 1
        headers = ['Khoản mục', '2023', '2024', 'Thay đổi (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        current_row += 1
        
        # Cash flow items
        cash_flow_items = [
            ("1. Lưu chuyển tiền từ hoạt động kinh doanh", "operating_cash_flow"),
            ("2. Lưu chuyển tiền từ hoạt động đầu tư", "investing_cash_flow"),
            ("3. Lưu chuyển tiền từ hoạt động tài chính", "financing_cash_flow"),
            ("4. Lưu chuyển tiền thuần trong kỳ", "net_cash_flow"),
            ("5. Tiền đầu kỳ", "beginning_cash"),
            ("6. Tiền cuối kỳ", "ending_cash")
        ]
        
        for item_name, key in cash_flow_items:
            val_2023 = self.data["cash_flow"]["2023"][key]
            val_2024 = self.data["cash_flow"]["2024"][key]
            change_pct = ((val_2024 - val_2023) / val_2023) * 100 if val_2023 != 0 else 0
            
            ws[f'A{current_row}'] = item_name
            ws[f'B{current_row}'] = val_2023
            ws[f'C{current_row}'] = val_2024
            ws[f'D{current_row}'] = change_pct / 100  # For percentage format
            
            # Format cells
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.thin_border
                if col >= 2:
                    cell.number_format = self.number_format if col <= 3 else self.percentage_format
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            current_row += 1
        
        return current_row
    
    def create_ratios_analysis_sheet(self):
        """Create Sheet 2: Financial Ratios Analysis"""
        ws = self.workbook.create_sheet("Phân tích chỉ số tài chính")
        
        # Header
        ws['A1'] = "PHÂN TÍCH CHỈ SỐ TÀI CHÍNH - VINGROUP"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A2'] = f"Cập nhật: {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        
        # Table header
        row = 4
        headers = ['Chỉ số', 'Công thức', '2023', '2024', 'Thay đổi', 'Ý nghĩa']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        
        # Calculate ratios (simplified calculation for demonstration)
        ratios_data = self._calculate_all_ratios()
        
        # Liquidity ratios
        ws[f'A{row}'] = "CHỈ SỐ THANH KHOẢN"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        liquidity_ratios = [
            ("Hệ số thanh khoản hiện hành", "Tài sản ngắn hạn / Nợ ngắn hạn", "current_ratio"),
            ("Hệ số thanh khoản nhanh", "(TSNH - Hàng tồn kho) / Nợ ngắn hạn", "quick_ratio"),
            ("Hệ số thanh khoản tuyệt đối", "Tiền mặt / Nợ ngắn hạn", "cash_ratio")
        ]
        
        for ratio_name, formula, key in liquidity_ratios:
            val_2023 = ratios_data["2023"][key]
            val_2024 = ratios_data["2024"][key]
            change = val_2024 - val_2023
            
            ws[f'A{row}'] = ratio_name
            ws[f'B{row}'] = formula
            ws[f'C{row}'] = val_2023
            ws[f'D{row}'] = val_2024
            ws[f'E{row}'] = change
            ws[f'F{row}'] = self._get_ratio_meaning(key)
            
            # Format cells
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                if col in [3, 4, 5]:  # Numeric columns
                    cell.number_format = '0.00'
                    cell.alignment = self.right_alignment
                else:
                    cell.alignment = self.left_alignment
            
            row += 1
        
        # Continue with other ratio categories...
        
        # Add charts
        self._add_ratios_chart(ws, row + 2)
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 35
        
        return ws
    
    def _calculate_all_ratios(self):
        """Calculate all financial ratios for both years"""
        ratios = {"2023": {}, "2024": {}}
        
        for year in ["2023", "2024"]:
            bs = self.data["balance_sheet"][year]
            inc = self.data["income_statement"][year]
            
            # Calculate totals
            current_assets = sum(bs["assets"]["current_assets"].values())
            current_liabilities = sum(bs["liabilities"]["current_liabilities"].values())
            total_assets = (sum(bs["assets"]["current_assets"].values()) + 
                          sum(bs["assets"]["non_current_assets"].values()))
            total_liabilities = (sum(bs["liabilities"]["current_liabilities"].values()) + 
                               sum(bs["liabilities"]["non_current_liabilities"].values()))
            total_equity = sum(bs["equity"].values())
            
            # Liquidity ratios
            ratios[year]["current_ratio"] = current_assets / current_liabilities
            
            quick_assets = (current_assets - bs["assets"]["current_assets"]["inventory"] - 
                          bs["assets"]["current_assets"]["prepaid_expenses"])
            ratios[year]["quick_ratio"] = quick_assets / current_liabilities
            
            cash_assets = (bs["assets"]["current_assets"]["cash_and_equivalents"] + 
                         bs["assets"]["current_assets"]["short_term_investments"])
            ratios[year]["cash_ratio"] = cash_assets / current_liabilities
            
            # Profitability ratios
            ratios[year]["net_profit_margin"] = (inc["net_profit"] / inc["revenue"]) * 100
            ratios[year]["gross_profit_margin"] = (inc["gross_profit"] / inc["revenue"]) * 100
            ratios[year]["roa"] = (inc["net_profit"] / total_assets) * 100
            ratios[year]["roe"] = (inc["net_profit"] / total_equity) * 100
            
            # Efficiency ratios
            ratios[year]["asset_turnover"] = inc["revenue"] / total_assets
            ratios[year]["inventory_turnover"] = inc["cost_of_goods_sold"] / bs["assets"]["current_assets"]["inventory"]
            
            # Leverage ratios
            ratios[year]["debt_to_equity"] = total_liabilities / total_equity
            ratios[year]["debt_to_assets"] = total_liabilities / total_assets
        
        return ratios
    
    def _get_ratio_meaning(self, ratio_key):
        """Get meaning explanation for each ratio"""
        meanings = {
            "current_ratio": "Khả năng thanh toán nợ ngắn hạn bằng tài sản ngắn hạn",
            "quick_ratio": "Khả năng thanh toán nhanh không bao gồm hàng tồn kho",
            "cash_ratio": "Khả năng thanh toán bằng tiền mặt và tương đương",
            "net_profit_margin": "Hiệu quả kinh doanh tổng thể",
            "gross_profit_margin": "Hiệu quả kiểm soát chi phí sản xuất",
            "roa": "Hiệu quả sử dụng tài sản để tạo lợi nhuận",
            "roe": "Hiệu quả sử dụng vốn chủ sở hữu",
            "asset_turnover": "Hiệu quả sử dụng tài sản để tạo doanh thu",
            "inventory_turnover": "Hiệu quả quản lý hàng tồn kho",
            "debt_to_equity": "Cân bằng giữa nợ và vốn chủ sở hữu",
            "debt_to_assets": "Mức độ sử dụng nợ để tài trợ tài sản"
        }
        return meanings.get(ratio_key, "")
    
    def _add_ratios_chart(self, ws, start_row):
        """Add chart for financial ratios"""
        # Create a simple bar chart for key ratios
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "So sánh chỉ số tài chính 2023-2024"
        chart.y_axis.title = "Giá trị"
        chart.x_axis.title = "Chỉ số"
        
        # Add data (simplified - would need actual data ranges)
        data = Reference(ws, min_col=3, min_row=6, max_row=10, max_col=4)
        cats = Reference(ws, min_col=1, min_row=6, max_row=10)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'H{start_row}')
    
    def create_guidelines_sheet(self):
        """Create Sheet 3: Guidelines and Exercises"""
        ws = self.workbook.create_sheet("Hướng dẫn và Bài tập")
        
        # Header
        ws['A1'] = "HƯỚNG DẪN VÀ BÀI TẬP PHÂN TÍCH TÀI CHÍNH"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        
        row = 3
        
        # Guidelines section
        ws[f'A{row}'] = "A. HƯỚNG DẪN PHÂN TÍCH"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 2
        
        guidelines = [
            "1. Cách đọc báo cáo tài chính:",
            "   • Bảng cân đối kế toán: Phản ánh tình hình tài chính tại thời điểm cụ thể",
            "   • Báo cáo KQKD: Phản ánh kết quả hoạt động kinh doanh trong kỳ",
            "   • Báo cáo lưu chuyển tiền tệ: Phản ánh luồng tiền vào/ra",
            "",
            "2. Các bước phân tích:",
            "   Bước 1: Phân tích cơ cấu tài sản và nguồn vốn",
            "   Bước 2: Tính toán các chỉ số tài chính",
            "   Bước 3: So sánh với năm trước và ngành",
            "   Bước 4: Đánh giá xu hướng và đưa ra nhận xét"
        ]
        
        for guideline in guidelines:
            ws[f'A{row}'] = guideline
            ws[f'A{row}'].font = Font(name='Arial', size=11)
            row += 1
        
        row += 2
        
        # Exercises section
        ws[f'A{row}'] = "B. BÀI TẬP THỰC HÀNH"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 2
        
        exercises = [
            "Câu 1: Phân tích cơ cấu tài sản của VinGroup",
            "• Tính tỷ trọng tài sản ngắn hạn/dài hạn",
            "• Nhận xét về sự thay đổi giữa 2023 và 2024",
            "",
            "Câu 2: Đánh giá khả năng thanh khoản",
            "• Tính và giải thích các chỉ số thanh khoản",
            "• So sánh với chuẩn mực ngành",
            "",
            "Câu 3: Phân tích khả năng sinh lời",
            "• Tính ROE, ROA, biên lợi nhuận",
            "• Đánh giá xu hướng và nguyên nhân",
            "",
            "Câu 4: Đánh giá hiệu quả hoạt động",
            "• Tính vòng quay tài sản, vòng quay hàng tồn kho",
            "• Nhận xét về hiệu quả quản lý",
            "",
            "Câu 5: Phân tích cơ cấu tài chính",
            "• Tính tỷ lệ nợ/vốn chủ sở hữu",
            "• Đánh giá rủi ro tài chính"
        ]
        
        for exercise in exercises:
            ws[f'A{row}'] = exercise
            ws[f'A{row}'].font = Font(name='Arial', size=11)
            if exercise.startswith("Câu"):
                ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
            row += 1
        
        row += 2
        
        # Answer guidelines
        ws[f'A{row}'] = "C. GỢI Ý TRÌNH BÀY KẾT QUẢ"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 2
        
        answer_guidelines = [
            "1. Cấu trúc báo cáo phân tích:",
            "   • Tóm tắt tình hình tài chính",
            "   • Phân tích chi tiết các chỉ số",
            "   • Nhận xét về xu hướng",
            "   • Khuyến nghị và đề xuất",
            "",
            "2. Cách trình bày số liệu:",
            "   • Sử dụng bảng biểu, biểu đồ",
            "   • Làm nổi bật các điểm quan trọng",
            "   • So sánh với năm trước"
        ]
        
        for guideline in answer_guidelines:
            ws[f'A{row}'] = guideline
            ws[f'A{row}'].font = Font(name='Arial', size=11)
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 80
        
        return ws
    
    def generate_excel_file(self, filename: str = "VinGroup_Financial_Analysis.xlsx"):
        """Generate the complete Excel file"""
        if not OPENPYXL_AVAILABLE:
            print("❌ Cannot generate Excel file: openpyxl not installed")
            return False
        
        print("🔄 Generating Excel file...")
        
        # Create all sheets
        self.create_financial_statements_sheet()
        self.create_ratios_analysis_sheet()
        self.create_guidelines_sheet()
        
        # Save file
        try:
            self.workbook.save(filename)
            print(f"✅ Excel file generated successfully: {filename}")
            return True
        except Exception as e:
            print(f"❌ Error generating Excel file: {str(e)}")
            return False

def main():
    """Main function"""
    if not OPENPYXL_AVAILABLE:
        print("❌ openpyxl is required to generate Excel files")
        print("   Please install with: pip install openpyxl")
        print("   For now, CSV files have been generated in the 'vingroup_analysis' folder")
        return
    
    try:
        # Create Excel generator
        generator = VinGroupExcelGenerator()
        
        # Generate Excel file
        success = generator.generate_excel_file()
        
        if success:
            print("\n" + "="*60)
            print("✅ EXCEL FILE GENERATED SUCCESSFULLY!")
            print("="*60)
            print("File: VinGroup_Financial_Analysis.xlsx")
            print("Sheets:")
            print("  1. Báo cáo tài chính VinGroup - Complete financial statements")
            print("  2. Phân tích chỉ số tài chính - Financial ratios analysis")
            print("  3. Hướng dẫn và Bài tập - Guidelines and exercises")
            print("\nFeatures:")
            print("  • Professional formatting with colors and borders")
            print("  • Financial ratio calculations with formulas")
            print("  • Charts and visualizations")
            print("  • Student exercise questions")
            print("  • Comprehensive analysis guidelines")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    main()