"""
Formula Validator - Hệ thống Kiểm tra và Xác thực Công thức
==========================================================

Kiểm tra tính chính xác của công thức Excel, named ranges và 
phương trình cân đối kế toán (Assets = Liabilities + Equity)

Tác giả: Hệ thống Phân tích Tài chính Động
Chuẩn: VAS/Circular 200/2014/TT-BTC
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
import datetime
import json
import os
import shutil
from pathlib import Path

class FormulaValidator:
    def __init__(self, excel_file=None):
        self.excel_file = excel_file
        self.wb = None
        self.validation_results = {
            'timestamp': datetime.datetime.now().isoformat(),
            'file_name': excel_file,
            'balance_equation': {},
            'named_ranges': {},
            'formula_errors': [],
            'data_validation': {},
            'overall_status': 'unknown'
        }
        
    def load_workbook(self, filename=None):
        """Tải workbook Excel để kiểm tra"""
        if filename:
            self.excel_file = filename
            
        if not self.excel_file or not os.path.exists(self.excel_file):
            print(f"❌ Không tìm thấy file: {self.excel_file}")
            return False
            
        try:
            self.wb = openpyxl.load_workbook(self.excel_file, data_only=False)
            print(f"✅ Đã tải workbook: {self.excel_file}")
            return True
        except Exception as e:
            print(f"❌ Lỗi tải workbook: {e}")
            return False
            
    def create_backup(self):
        """Tạo backup file trước khi validation"""
        if not self.excel_file:
            return None
            
        try:
            # Tạo thư mục backup nếu chưa tồn tại
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            
            # Tạo tên file backup
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = Path(self.excel_file).stem
            extension = Path(self.excel_file).suffix
            backup_filename = f"{backup_dir}/{filename}_backup_{timestamp}{extension}"
            
            # Copy file
            shutil.copy2(self.excel_file, backup_filename)
            print(f"✅ Đã tạo backup: {backup_filename}")
            
            self.validation_results['backup_file'] = backup_filename
            return backup_filename
            
        except Exception as e:
            print(f"⚠️  Lỗi tạo backup: {e}")
            return None
            
    def validate_balance_equation(self):
        """Kiểm tra phương trình cân đối: Assets = Liabilities + Equity"""
        print("⚖️  Đang kiểm tra phương trình cân đối kế toán...")
        
        try:
            # Tìm sheet bảng cân đối kế toán
            balance_sheet = None
            for sheet_name in self.wb.sheetnames:
                if 'cân đối' in sheet_name.lower() or 'balance' in sheet_name.lower():
                    balance_sheet = self.wb[sheet_name]
                    break
                    
            if not balance_sheet:
                balance_sheet = self.wb.worksheets[0]  # Dùng sheet đầu tiên
                
            print(f"📊 Đang kiểm tra sheet: {balance_sheet.title}")
            
            # Tìm các giá trị tổng
            total_assets = None
            total_liabilities = None
            total_equity = None
            total_liab_equity = None
            
            # Quét qua các cell để tìm tổng
            for row in balance_sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        
                        # Tìm tổng tài sản
                        if 'tổng' in cell_text and 'tài sản' in cell_text:
                            value_cell = balance_sheet.cell(row=cell.row, column=4)  # Cột D
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                total_assets = value_cell.value
                                print(f"   📈 Tổng tài sản: {total_assets:,.0f}")
                                
                        # Tìm tổng nợ
                        elif 'tổng' in cell_text and 'nợ' in cell_text:
                            value_cell = balance_sheet.cell(row=cell.row, column=4)
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                total_liabilities = value_cell.value
                                print(f"   📉 Tổng nợ phải trả: {total_liabilities:,.0f}")
                                
                        # Tìm tổng vốn chủ sở hữu
                        elif 'tổng' in cell_text and 'vốn chủ' in cell_text:
                            value_cell = balance_sheet.cell(row=cell.row, column=4)
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                total_equity = value_cell.value
                                print(f"   💰 Tổng vốn chủ sở hữu: {total_equity:,.0f}")
                                
                        # Tìm tổng nguồn vốn
                        elif 'tổng' in cell_text and 'nguồn vốn' in cell_text:
                            value_cell = balance_sheet.cell(row=cell.row, column=4)
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                total_liab_equity = value_cell.value
                                print(f"   🏦 Tổng nguồn vốn: {total_liab_equity:,.0f}")
                                
            # Kiểm tra cân đối
            balance_result = {
                'total_assets': total_assets,
                'total_liabilities': total_liabilities,
                'total_equity': total_equity,
                'total_liab_equity': total_liab_equity,
                'is_balanced': False,
                'difference': 0,
                'status': 'unknown'
            }
            
            if total_assets and total_liab_equity:
                balance_result['difference'] = abs(total_assets - total_liab_equity)
                balance_result['is_balanced'] = balance_result['difference'] < 1000  # Cho phép sai số nhỏ
                
                if balance_result['is_balanced']:
                    balance_result['status'] = 'balanced'
                    print("✅ Phương trình cân đối CHÍNH XÁC!")
                else:
                    balance_result['status'] = 'unbalanced'
                    print(f"❌ Phương trình cân đối KHÔNG CHÍNH XÁC!")
                    print(f"   📊 Chênh lệch: {balance_result['difference']:,.0f}")
                    
            # Kiểm tra cân đối phụ (nếu có tách riêng nợ và vốn)
            if total_liabilities and total_equity:
                calculated_total = total_liabilities + total_equity
                if total_assets:
                    alt_difference = abs(total_assets - calculated_total)
                    print(f"   🔍 Kiểm tra phụ: Assets vs (Liabilities + Equity)")
                    print(f"      Chênh lệch: {alt_difference:,.0f}")
                    
            self.validation_results['balance_equation'] = balance_result
            return balance_result['is_balanced']
            
        except Exception as e:
            print(f"❌ Lỗi kiểm tra phương trình cân đối: {e}")
            self.validation_results['balance_equation']['error'] = str(e)
            return False
            
    def validate_named_ranges(self):
        """Kiểm tra tính hợp lệ của named ranges"""
        print("🏷️  Đang kiểm tra named ranges...")
        
        try:
            named_ranges_info = {
                'total_count': 0,
                'valid_ranges': [],
                'invalid_ranges': [],
                'missing_references': [],
                'status': 'unknown'
            }
            
            # Kiểm tra named ranges trong workbook
            if hasattr(self.wb, 'defined_names') and self.wb.defined_names:
                named_ranges_info['total_count'] = len(self.wb.defined_names)
                print(f"📊 Tổng số named ranges: {named_ranges_info['total_count']}")
                
                for name, defined_name in self.wb.defined_names.items():
                    try:
                        # Kiểm tra reference có hợp lệ không
                        if hasattr(defined_name, 'attr_text') and defined_name.attr_text:
                            # Kiểm tra cell reference
                            ref = defined_name.attr_text
                            is_valid = self.validate_cell_reference(ref)
                            
                            range_info = {
                                'name': name,
                                'reference': ref,
                                'is_valid': is_valid
                            }
                            
                            if is_valid:
                                named_ranges_info['valid_ranges'].append(range_info)
                                print(f"   ✅ {name}: {ref}")
                            else:
                                named_ranges_info['invalid_ranges'].append(range_info)
                                print(f"   ❌ {name}: {ref} (không hợp lệ)")
                                
                        else:
                            named_ranges_info['invalid_ranges'].append({
                                'name': name,
                                'reference': 'undefined',
                                'is_valid': False
                            })
                            print(f"   ❌ {name}: Không có reference")
                            
                    except Exception as e:
                        print(f"   ⚠️  Lỗi kiểm tra {name}: {e}")
                        named_ranges_info['invalid_ranges'].append({
                            'name': name,
                            'reference': 'error',
                            'is_valid': False,
                            'error': str(e)
                        })
                        
            else:
                print("⚠️  Không tìm thấy named ranges trong workbook")
                
            # Kiểm tra các named ranges cần thiết
            required_ranges = [
                'TotalAssets', 'CurrentAssets', 'NonCurrentAssets',
                'TotalLiabilities', 'CurrentLiabilities', 'NonCurrentLiabilities',
                'TotalEquity', 'Revenue', 'NetIncome'
            ]
            
            existing_names = [r['name'] for r in named_ranges_info['valid_ranges']]
            for required in required_ranges:
                if required not in existing_names:
                    named_ranges_info['missing_references'].append(required)
                    
            if named_ranges_info['missing_references']:
                print(f"⚠️  Thiếu named ranges: {', '.join(named_ranges_info['missing_references'])}")
                
            # Đánh giá tổng thể
            valid_count = len(named_ranges_info['valid_ranges'])
            total_count = named_ranges_info['total_count']
            missing_count = len(named_ranges_info['missing_references'])
            
            if total_count > 0 and valid_count / total_count >= 0.8 and missing_count == 0:
                named_ranges_info['status'] = 'good'
                print("✅ Named ranges: TỐT")
            elif valid_count > 0:
                named_ranges_info['status'] = 'warning'
                print("⚠️  Named ranges: CẦN CHÚ Ý")
            else:
                named_ranges_info['status'] = 'poor'
                print("❌ Named ranges: THIẾU HOẶC LỖI")
                
            self.validation_results['named_ranges'] = named_ranges_info
            return named_ranges_info['status'] in ['good', 'warning']
            
        except Exception as e:
            print(f"❌ Lỗi kiểm tra named ranges: {e}")
            self.validation_results['named_ranges']['error'] = str(e)
            return False
            
    def validate_cell_reference(self, reference):
        """Kiểm tra tính hợp lệ của cell reference"""
        try:
            # Kiểm tra format cơ bản của cell reference
            if not reference:
                return False
                
            # Loại bỏ sheet name nếu có
            if '!' in reference:
                sheet_part, cell_part = reference.split('!', 1)
                # Kiểm tra sheet có tồn tại không
                sheet_name = sheet_part.strip("'\"")
                if sheet_name not in self.wb.sheetnames:
                    return False
                reference = cell_part
                
            # Kiểm tra format cell (như D10, A1:B5)
            import re
            cell_pattern = r'^[A-Z]+\d+$|^[A-Z]+\d+:[A-Z]+\d+$'
            return bool(re.match(cell_pattern, reference.strip()))
            
        except:
            return False
            
    def validate_formulas(self):
        """Kiểm tra công thức Excel có lỗi không"""
        print("🧮 Đang kiểm tra công thức Excel...")
        
        formula_errors = []
        
        try:
            for sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                print(f"   📋 Đang kiểm tra sheet: {sheet_name}")
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula = cell.value
                            
                            # Kiểm tra các lỗi phổ biến
                            error_checks = [
                                ('#REF!', 'Reference error - Tham chiếu không hợp lệ'),
                                ('#DIV/0!', 'Division by zero - Chia cho 0'),
                                ('#NAME?', 'Name error - Tên không được nhận dạng'),
                                ('#VALUE!', 'Value error - Lỗi giá trị'),
                                ('#N/A', 'Not available - Giá trị không có'),
                                ('#NULL!', 'Null error - Lỗi null'),
                                ('#NUM!', 'Number error - Lỗi số')
                            ]
                            
                            for error_code, description in error_checks:
                                if error_code in formula:
                                    error_info = {
                                        'sheet': sheet_name,
                                        'cell': f"{cell.column_letter}{cell.row}",
                                        'formula': formula,
                                        'error_type': error_code,
                                        'description': description
                                    }
                                    formula_errors.append(error_info)
                                    print(f"   ❌ {error_info['cell']}: {error_code} - {description}")
                                    
            if not formula_errors:
                print("✅ Không tìm thấy lỗi công thức")
            else:
                print(f"⚠️  Tìm thấy {len(formula_errors)} lỗi công thức")
                
            self.validation_results['formula_errors'] = formula_errors
            return len(formula_errors) == 0
            
        except Exception as e:
            print(f"❌ Lỗi kiểm tra công thức: {e}")
            return False
            
    def validate_data_integrity(self):
        """Kiểm tra tính toàn vẹn dữ liệu"""
        print("🔍 Đang kiểm tra tính toàn vẹn dữ liệu...")
        
        try:
            data_validation = {
                'positive_values': True,
                'reasonable_ranges': True,
                'data_types': True,
                'completeness': True,
                'issues': []
            }
            
            # Tìm sheet bảng cân đối
            balance_sheet = None
            for sheet_name in self.wb.sheetnames:
                if 'cân đối' in sheet_name.lower() or 'balance' in sheet_name.lower():
                    balance_sheet = self.wb[sheet_name]
                    break
                    
            if balance_sheet:
                print(f"   📊 Kiểm tra dữ liệu trong sheet: {balance_sheet.title}")
                
                # Kiểm tra các khoản mục chính phải có giá trị dương
                positive_check_items = [
                    'tài sản', 'vốn chủ sở hữu', 'tiền', 'hàng tồn kho'
                ]
                
                for row in balance_sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = cell.value.lower()
                            
                            for item in positive_check_items:
                                if item in cell_text and 'tổng' not in cell_text:
                                    # Kiểm tra giá trị tương ứng
                                    value_cell = balance_sheet.cell(row=cell.row, column=4)
                                    if value_cell.value and isinstance(value_cell.value, (int, float)):
                                        if value_cell.value < 0:
                                            issue = f"Giá trị âm tại {cell.coordinate}: {cell_text}"
                                            data_validation['issues'].append(issue)
                                            data_validation['positive_values'] = False
                                            print(f"   ⚠️  {issue}")
                                            
                # Kiểm tra khoảng giá trị hợp lý (không quá lớn hoặc quá nhỏ)
                for row in balance_sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, (int, float)):
                            if cell.value > 1e15:  # Quá lớn
                                issue = f"Giá trị quá lớn tại {cell.coordinate}: {cell.value}"
                                data_validation['issues'].append(issue)
                                data_validation['reasonable_ranges'] = False
                                print(f"   ⚠️  {issue}")
                            elif cell.value < -1e15:  # Quá nhỏ (âm)
                                issue = f"Giá trị quá nhỏ tại {cell.coordinate}: {cell.value}"
                                data_validation['issues'].append(issue)
                                data_validation['reasonable_ranges'] = False
                                print(f"   ⚠️  {issue}")
                                
            if not data_validation['issues']:
                print("✅ Dữ liệu toàn vẹn")
            else:
                print(f"⚠️  Tìm thấy {len(data_validation['issues'])} vấn đề dữ liệu")
                
            self.validation_results['data_validation'] = data_validation
            return len(data_validation['issues']) == 0
            
        except Exception as e:
            print(f"❌ Lỗi kiểm tra dữ liệu: {e}")
            return False
            
    def run_complete_validation(self, create_backup=True):
        """Chạy toàn bộ quá trình validation"""
        print("🔍 BẮT ĐẦU VALIDATION TOÀN DIỆN...")
        print("=" * 60)
        
        # Tạo backup nếu cần
        if create_backup:
            self.create_backup()
            
        # Load workbook
        if not self.load_workbook():
            return False
            
        # Chạy các kiểm tra
        checks = {
            'balance_equation': self.validate_balance_equation(),
            'named_ranges': self.validate_named_ranges(), 
            'formulas': self.validate_formulas(),
            'data_integrity': self.validate_data_integrity()
        }
        
        # Đánh giá tổng thể
        passed_checks = sum(checks.values())
        total_checks = len(checks)
        
        if passed_checks == total_checks:
            self.validation_results['overall_status'] = 'excellent'
            status_msg = "XUẤT SẮC"
            status_icon = "🎉"
        elif passed_checks >= total_checks * 0.75:
            self.validation_results['overall_status'] = 'good'
            status_msg = "TỐT"
            status_icon = "✅"
        elif passed_checks >= total_checks * 0.5:
            self.validation_results['overall_status'] = 'fair'
            status_msg = "CHẤP NHẬN ĐƯỢC"
            status_icon = "⚠️"
        else:
            self.validation_results['overall_status'] = 'poor'
            status_msg = "CẦN CẢI THIỆN"
            status_icon = "❌"
            
        print("\n" + "=" * 60)
        print(f"{status_icon} KẾT QUẢ VALIDATION TỔNG THỂ: {status_msg}")
        print(f"📊 Đã qua: {passed_checks}/{total_checks} kiểm tra")
        
        # In chi tiết kết quả
        print("\n📋 CHI TIẾT KẾT QUẢ:")
        for check_name, result in checks.items():
            icon = "✅" if result else "❌"
            name_map = {
                'balance_equation': 'Phương trình cân đối',
                'named_ranges': 'Named ranges', 
                'formulas': 'Công thức Excel',
                'data_integrity': 'Tính toàn vẹn dữ liệu'
            }
            print(f"   {icon} {name_map.get(check_name, check_name)}: {'PASS' if result else 'FAIL'}")
            
        # Lưu kết quả
        self.save_validation_report()
        
        return self.validation_results['overall_status'] in ['excellent', 'good']
        
    def save_validation_report(self):
        """Lưu báo cáo validation"""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Lưu JSON report
            json_filename = f"validation_report_{timestamp}.json"
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(self.validation_results, f, ensure_ascii=False, indent=2)
            print(f"✅ Đã lưu báo cáo JSON: {json_filename}")
            
            # Lưu text report
            text_filename = f"validation_report_{timestamp}.txt"
            with open(text_filename, 'w', encoding='utf-8') as f:
                f.write("BÁO CÁO VALIDATION HỆ THỐNG TÀI CHÍNH ĐỘNG\n")
                f.write("=" * 60 + "\n")
                f.write(f"Thời gian: {self.validation_results['timestamp']}\n")
                f.write(f"File: {self.validation_results['file_name']}\n")
                f.write(f"Trạng thái tổng thể: {self.validation_results['overall_status']}\n\n")
                
                # Chi tiết phương trình cân đối
                if 'balance_equation' in self.validation_results:
                    f.write("1. PHƯƠNG TRÌNH CÂN ĐỐI KẾ TOÁN:\n")
                    balance = self.validation_results['balance_equation']
                    f.write(f"   - Tổng tài sản: {balance.get('total_assets', 'N/A')}\n")
                    f.write(f"   - Tổng nguồn vốn: {balance.get('total_liab_equity', 'N/A')}\n")
                    f.write(f"   - Chênh lệch: {balance.get('difference', 'N/A')}\n")
                    f.write(f"   - Trạng thái: {balance.get('status', 'N/A')}\n\n")
                    
                # Chi tiết named ranges
                if 'named_ranges' in self.validation_results:
                    f.write("2. NAMED RANGES:\n")
                    nr = self.validation_results['named_ranges']
                    f.write(f"   - Tổng số: {nr.get('total_count', 0)}\n")
                    f.write(f"   - Hợp lệ: {len(nr.get('valid_ranges', []))}\n")
                    f.write(f"   - Không hợp lệ: {len(nr.get('invalid_ranges', []))}\n")
                    f.write(f"   - Thiếu: {len(nr.get('missing_references', []))}\n\n")
                    
                # Chi tiết lỗi công thức
                if 'formula_errors' in self.validation_results:
                    f.write("3. LỖI CÔNG THỨC:\n")
                    errors = self.validation_results['formula_errors']
                    f.write(f"   - Số lỗi: {len(errors)}\n")
                    for error in errors:
                        f.write(f"   - {error.get('sheet')}.{error.get('cell')}: {error.get('error_type')}\n")
                    f.write("\n")
                    
                # Chi tiết dữ liệu
                if 'data_validation' in self.validation_results:
                    f.write("4. TÍNH TOÀN VẸN DỮ LIỆU:\n")
                    data = self.validation_results['data_validation']
                    f.write(f"   - Giá trị dương: {data.get('positive_values', False)}\n")
                    f.write(f"   - Khoảng hợp lý: {data.get('reasonable_ranges', False)}\n")
                    f.write(f"   - Số vấn đề: {len(data.get('issues', []))}\n")
                    
            print(f"✅ Đã lưu báo cáo text: {text_filename}")
            
            return json_filename, text_filename
            
        except Exception as e:
            print(f"⚠️  Lỗi lưu báo cáo: {e}")
            return None, None

# Test và chạy
if __name__ == "__main__":
    # Kiểm tra file balance sheet mới tạo
    balance_files = [f for f in os.listdir('.') if f.startswith('bang_can_doi_ke_toan_dynamic_') and f.endswith('.xlsx')]
    
    if balance_files:
        latest_file = max(balance_files, key=os.path.getctime)
        print(f"🔍 Tìm thấy file balance sheet: {latest_file}")
        
        validator = FormulaValidator(latest_file)
        result = validator.run_complete_validation()
        
        if result:
            print("\n🎉 HỆ THỐNG ĐÃ ĐẬT CHUẨN VALIDATION!")
        else:
            print("\n⚠️  HỆ THỐNG CẦN CẢI THIỆN!")
            
        print("\n💡 Khuyến nghị:")
        print("1. Kiểm tra báo cáo validation chi tiết")
        print("2. Sửa các lỗi được phát hiện")
        print("3. Chạy lại validation sau khi sửa")
        print("4. Backup file trước khi thay đổi")
        
    else:
        print("❌ Không tìm thấy file balance sheet để kiểm tra")
        print("💡 Chạy enhanced_balance_sheet_generator.py trước")